# coding: utf-8

import requests
from bs4 import BeautifulSoup
import re
import tqdm
import fiscal_advisors as fa
import xlwt
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support.ui import WebDriverWait as Wait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support.wait import WebDriverWait
from selenium.webdriver.support.ui import Select
from selenium.webdriver.support import expected_conditions as Expect
from selenium.webdriver.chrome.options import Options
import os
import openpyxl


def get_results_page_to_local(fromDate, toDate, infoPageName, textPath):
    """
    get all basic information from the results page, save it to a tsv file; get all the summary pages, and save them to local data path
    :param fromDate: start date for search reaults
    :param toDate: end date for search reaults
    :param infoPageName: output tsv file name
    :param textPath: output text file path/dir
    """
    # load the search page
    url1 = 'https://data.bondbuyer.com/salesresults/SalesResult/GetSalesDetails?&FromDate={}&ToDate={}'.format(
        fromDate, toDate)
    chrome_options = Options()
    chrome_options.add_argument('--headless')
    browser = webdriver.Chrome(options=chrome_options)
    browser.get(url1)
    # a little trick, load again to get all results
    url2 = 'https://data.bondbuyer.com/salesresults/SalesResult/GetSalesDetails?page=1&pagesize=10000&FromDate={}&ToDate={}'.format(
        fromDate, toDate)
    browser.get(url2)
    # wait a moment, until it loaded all needed information
    Wait(browser, 600).until(
        Expect.presence_of_element_located(
            (By.CSS_SELECTOR, "body>table>tbody>tr"))
    )
    trs = browser.find_elements_by_css_selector("body>table>tbody>tr")
    # get all needed information
    if len(trs) > 1:
        with open(infoPageName, 'w') as iP:
            par = tqdm.tqdm(total=len(trs), ncols=80)
            count = 0
            iP.write('ID' + '\t'+'Headline' + '\t' +
                     'Summary link' + '\t' + 'Publish Date'+'\n')
            for tr in trs:
                par.update(1)
                trId = tr.get_attribute('id')
                if not str(trId).startswith('dealHeader'):
                    a = tr.find_element_by_css_selector('a')
                    summaryLink = a.get_attribute('href')
                    Id = summaryLink.split('GetDetails/')[1]
                    title = a.text.lower()
                    title = title.replace('results', 'Results')
                    date = tr.find_element_by_css_selector('span').text
                    # get summary page and save to local
                    page = get_page(summaryLink)
                    if page is not None:
                        title = get_details(textPath, page)
                    iP.write(Id + '\t'+title + '\t' +
                             summaryLink + '\t' + date+'\n')
                    count += 1
        iP.close()
    print('total page is {}'.format(str(count)))


def gen_url(startUrl, id):
    # get the full url
    return startUrl+str(id)


def get_page(url):
    """
    get web source code from the summary webpages
    :param url: the summary or term webpage
    :return page: string data type, page source code
    """
    webPage = requests.get(url)
    if webPage.status_code == 404:
        return None
    else:
        page = webPage.text
        page = page.replace('</br>', '\n')
        return page


def get_id(tagId):
    """
    get ID of the TagId string code
    :param tagId: TagId string code
    :return Id: string data type, result id
    """
    Id = re.search(r'\d+', tagId)
    if Id is not None:
        return Id.group(0)


def get_all_ids(results):
    """
    get all the ids of the web source code
    :param results: a list of find results
    :return Ids: List data type, all results id
    """
    Ids = []
    for result in results:
        tagId = result['id']
        Ids.append(get_id(tagId))
    return Ids


def get_head(webPage):
    """
    get web page title
    :param webPage: webpage source code
    :return head: string data type, webpage title
    """
    soup = BeautifulSoup(webPage, 'lxml')
    head = soup.find('p', 'Headlinecls').get_text()
    return head


def get_details(filePath, webPage):
    '''
    parse the webpage souce and save the text data to local
    :param filePath: output raw text file path
    :param webPage: source code of the webpage
    '''
    soup = BeautifulSoup(webPage, 'lxml')
    head = soup.find('p', 'Headlinecls').get_text()
    if head is not None and head is not '':
        fileName = filePath + head+'.txt'
        with open(fileName, 'w') as out:
            salesResults = soup.find_all('p', 'RCSalesResultcls')
            # if the webpage contains the RCSalesResultcls, only save the rcs/tb/foot part of the page
            if len(salesResults) > 0:
                tblExcel = soup.find_all('table', 'tblExcel')
                tblExcelHide = soup.find_all('table', 'tblExcelHide')
                footRCSalesResult = soup.find_all('p', 'FootRCSalesResultcls')

                srIds = get_all_ids(salesResults)
                teIds = get_all_ids(tblExcel)
                tehIds = get_all_ids(tblExcelHide)
                frIds = get_all_ids(footRCSalesResult)
                for id in srIds:
                    if id in teIds and id in frIds:
                        srP = srIds.index(id)
                        teP = teIds.index(id)
                        frP = frIds.index(id)
                        out.write('======RCSalesResultcls======\n')
                        out.write(salesResults[srP].get_text(separator='\n'))
                        out.write('\n')
                        out.write('======tblExcel======\n')
                        out.write(tblExcel[teP].get_text(separator='\n'))
                        out.write('\n')
                        out.write('======FootRCSalesResultcls======\n')
                        out.write(
                            footRCSalesResult[frP].get_text(separator='\n'))
                        out.write('\n')
                    elif id in tehIds and id in frIds:
                        srP = srIds.index(id)
                        teP = tehIds.index(id)
                        frP = frIds.index(id)
                        out.write('======RCSalesResultcls======\n')
                        out.write(salesResults[srP].get_text(separator='\n'))
                        out.write('\n')
                        out.write('======tblExcel======\n')
                        out.write(tblExcelHide[teP].get_text(separator='\n'))
                        out.write('\n')
                        out.write('======FootRCSalesResultcls======\n')
                        out.write(
                            footRCSalesResult[frP].get_text(separator='\n'))
                        out.write('\n')
                    else:
                        print('index error')
            else:
                # if there is no RCSalesResultcls, save all source text
                productDataPrev = soup.select(
                    '#productDataPrev')[0].get_text(separator='\n')
                out.write(productDataPrev)
        out.close()
    return head


def get_webpage_to_local(startUrl, path, infoPageName):
    '''
    another way to get all raw data by using result id directly
    :param startUrl: url of the summary page
    :param path: output text file path/dir
    :param infoPageName: output tsv file name
    '''
    count = 0
    startId = 9400
    par = tqdm.tqdm(total=startId, ncols=80)
    Id = startId
    with open(infoPageName, 'w') as iP:
        iP.write('ID' + '\t'+'Page Name' + '\t' + 'Summary link' + '\n')
        while Id > 0:
            par.update(1)
            url = gen_url(startUrl, Id)
            page = get_page(url)
            if page is not None:
                head = get_details(path, page)
                if head is not None and head is not '':
                    count += 1
                    iP.write(str(Id) + '\t' + head + '\t' + url + '\n')
            Id -= 1
    iP.close()
    print('total page is {}'.format(str(count)))


def init_resule_value():
    """
    initialize the result dict
    :return allValue: dict like data type, a initialized data structure
    """
    allValue = dict()
    allValue['state'] = ''
    allValue['issuer'] = ''
    allValue['date'] = ''
    allValue['principal'] = ''
    allValue['description'] = ''
    allValue['dated'] = ''
    allValue['due'] = ''
    allValue['firstCoupon'] = ''
    allValue['callable'] = ''
    allValue['winning'] = ''
    allValue['purchased'] = ''
    allValue['form'] = ''
    allValue['lo'] = ''
    allValue['fa'] = ''
    allValue['otherBidder'] = ''
    allValue['otherManager'] = ''
    return allValue


def from_stream_to_list(webPage):
    """
    trun the raw text to a list
    :param webPage: text raw data of the page
    :return pageList: list like data type, a list of string which contain all text of the page
    """
    pageList = []
    pages = webPage.split('\n')
    for line in pages:
        line = line.strip()
        line = line.replace('&amp', '&')
        line = line.replace('&nbsp', ' ')
        if line is not '':
            pageList.append(line)
    return pageList


def get_result_pattern1(page):
    """
    trun the raw text to a structured results
    :param page: text raw data of the page
    :return results: list like data type, the structured reuslts
    """
    """
    example pattern 1: contains the RCSalesResultcls
    https://data.bondbuyer.com/salesresults/GetDetails/9378
    """
    results = []
    page = from_stream_to_list(page)
    allValue = init_resule_value()
    state = ''
    nonVoid = False
    for currentIndex, line in enumerate(page):
        if line == '======RCSalesResultcls======':
            if currentIndex > 1 and nonVoid:
                results.append(allValue)
            allValue = init_resule_value()
            nonVoid = False
            if not page[currentIndex+1].startswith('====') and not page[currentIndex+1].startswith('TAXABLE'):
                nonVoid = True
        if line in usState:
            state = line
        if re.match(r'\d*-(.*)-\d*(.*)\$(.*)', line):
            issuer = page[currentIndex-1]
            allValue['issuer'] = issuer
            c = line.split()
            date = c[0]
            principal = c[1]
            allValue['date'] = date
            allValue['principal'] = principal
            description = page[currentIndex+1]
            allValue['description'] = description
        if re.match(r'(.*) \d*, \d*', line) and page[currentIndex+1].isdigit():
            issuer = page[currentIndex-1]
            allValue['issuer'] = issuer
            date = line
            principal = page[currentIndex+1]
            allValue['date'] = date
            allValue['principal'] = '$' + principal
            description = page[currentIndex+2]
            allValue['description'] = description
        if 'Bonds' in line or 'Series' in line or 'Note' in line and allValue['description'] is not '':
            description = line
            allValue['description'] = description
        if line.startswith('Dated '):
            dated = line
            allValue['dated'] = dated
        if line.startswith('Due '):
            due = line
            allValue['due'] = due
        if line.startswith('First coupon'):
            firstCoupon = line
            allValue['firstCoupon'] = firstCoupon
        if 'Callable' in line or line.startswith('Non-callable'):
            Callable = line
            allValue['callable'] = Callable
        if line.startswith('Winning'):
            winning = line
            allValue['winning'] = winning
        if line.startswith('Purchased'):
            purchased = line
            allValue['purchased'] = purchased
        if line.startswith('L.O.:'):
            lo = line
            allValue['lo'] = lo
        if line.startswith('F.A.:'):
            fa = line
            allValue['fa'] = fa
        if line.startswith('Other bidders'):
            otherBidder = line
            i = currentIndex + 1
            while i < len(page)-1:
                if page[i] == '======RCSalesResultcls======' or page[i] == '======tblExcel======':
                    break
                otherBidder = otherBidder + '\n' + page[i]
                i += 1
            allValue['otherBidder'] = otherBidder
        if line.startswith('Other managers:'):
            otherManager = line
            i = currentIndex + 1
            while page[i].startswith('L.O.:') is False and page[i].startswith('F.A.:') is False:
                otherManager = otherManager + '\n' + page[i]
                i += 1
            allValue['otherManager'] = otherManager
        if line == 'DUE':
            i = currentIndex
            form = ''
            while page[i].startswith('======FootRCSalesResultcls======') is False:
                if re.match(r'\d+/\d+/\d+', page[i]):
                    form = form + '\n' + page[i]
                else:
                    form = form + '| ' + page[i]
                i += 1
            allValue['form'] = form
        allValue['state'] = state
        if currentIndex == len(page) - 1 and nonVoid:
            results.append(allValue)
    return results


def get_result_pattern2(page):
    """
    trun the raw text to a structured results
    :param page: text raw data of the page
    :return results: list like data type, the structured reuslts
    """
    """
    example pattern 2: non-containing of the RCSalesResultcls
    https://data.bondbuyer.com/salesresults/GetDetails/2
    """
    mon = ['Dec', 'Nov', 'Oct', 'Sep', 'Aug', 'Jul', 'Jun', 'May', 'Apr',
           'Mar', 'Feb', 'Jan']
    results = []
    page = from_stream_to_list(page)
    allValue = init_resule_value()
    state = ''
    date = ''
    principal = ''
    issuer = ''
    for currentIndex, line in enumerate(page):
        if re.match(r'(.*)\d*, \d* . . . . . . \$(.*)', line):
            c = line.split(' . . . . . . ')
            date = c[0]
            principal = c[1]
            issuer = page[currentIndex-1]
        if re.match(r'\d*-(.*)-\d*', line):
            issuer = page[currentIndex-1]
            c = line.split()
            date = c[0]
            if len(c) == 2:
                principal = c[1]
        if re.match(r'(.*)\ \d*,\ \d*(.*)\$(.*)', line):
            c = line.split()
            if c[0] in mon:
                issuer = page[currentIndex-1]
                date = c[0] + ' ' + c[1] +c[2]
                if len(c) == 4:
                    principal = c[3]
        if line in usState:
            state = line
        if 'Bonds' in line or 'Series' in line or 'Note' in line and (line is not 'California Education Notes Program') and (line is not 'California Communities Note Program'):
            if allValue['description'] is not '':
                results.append(allValue)
                allValue = init_resule_value()
            description = line
            allValue['description'] = description
            allValue['date'] = date
            allValue['principal'] = principal
            allValue['issuer'] = issuer
            allValue['state'] = state
        if line.startswith('Dated '):
            dated = line
            allValue['dated'] = dated
        if line.startswith('Due '):
            due = line
            allValue['due'] = due
        if line.startswith('First coupon'):
            firstCoupon = line
            allValue['firstCoupon'] = firstCoupon
        if 'Callable' in line or line.startswith('Non-callable'):
            Callable = line
            allValue['callable'] = Callable
        if line.startswith('Winning'):
            winning = line
            allValue['winning'] = winning
        if line.startswith('Purchased'):
            purchased = line
            allValue['purchased'] = purchased
        if line.startswith('L.O.:'):
            lo = line
            allValue['lo'] = lo
        if line.startswith('F.A.:'):
            fa = line
            allValue['fa'] = fa
        if line.startswith('Other bidders'):
            otherBidder = line
            i = currentIndex + 1
            while i < len(page)-1:
                if page[i] in usState or 'Bonds' in page[i] or 'Series' in page[i]  \
                        or 'Note' in page[i] or 'bank qualified' in page[i] or 'book entry' in page[i]:
                    break
                if i < len(page)-3 and ('Bonds' in page[i+2] or 'Series' in page[i+2]
                                        or 'Note' in page[i+2] or 'bank qualified' in page[i+2] or 'book entry' in page[i+2]):
                    break
                otherBidder = otherBidder + '\n' + page[i]
                i += 1
            allValue['otherBidder'] = otherBidder
        if line.startswith('Other managers:'):
            otherManager = line
            i = currentIndex + 1
            while i < len(page)-1:
                if page[i] in usState or 'Bonds' in page[i] or 'Series' in page[i]  \
                    or 'Note' in page[i] or 'bank qualified' in page[i] or 'book entry' in page[i]\
                        or page[i].startswith('L.O.:') or page[i].startswith('F.A.:') or page[i].startswith('Other bidders:'):
                    break
                if i < len(page)-3 and ('Bonds' in page[i+2] or 'Series' in page[i+2]
                                        or 'Note' in page[i+2] or 'bank qualified' in page[i+2] or 'book entry' in page[i+2]):
                    break
                otherManager = otherManager + '\n' + page[i]
                i += 1
            allValue['otherManager'] = otherManager
        if line == 'Due':
            i = currentIndex
            form = ''
            while i < len(page)-1:
                if page[i] in usState or 'Bonds' in page[i] or 'Series' in page[i]  \
                        or 'Note' in page[i] or 'bank qualified' in page[i] or 'book entry' in page[i]\
                    or page[i].startswith('L.O.:') or page[i].startswith('F.A.:') or \
                        page[i].startswith('Other managers:') or page[i].startswith('Other managers:'):
                    break
                if i < len(page)-3 and ('Bonds' in page[i+2] or 'Series' in page[i+2]
                                        or 'Note' in page[i+2] or 'bank qualified' in page[i+2] or 'book entry' in page[i+2]):
                    break
                if re.match(r'\d+/\d+/\d+', page[i]):
                    form = form + '\n' + page[i]
                else:
                    form = form + '| ' + page[i]
                i += 1
            allValue['form'] = form
        if currentIndex == len(page) - 1:
            results.append(allValue)
    return results


def get_pattern(page):
    """
    try to decode the infromation from all different patterns, and turn the raw data to a structured result
    :param page: the raw data of the page
    :return results: list like data type, the structured reuslts
    """
    results = []
    if '======RCSalesResultcls======' in page:
        results = get_result_pattern1(page)
    else:
        results = get_result_pattern2(page)
    return results


usState = ['ALABAMA', 'ALASKA', 'ARIZONA', 'ARKANSAS', 'CALIFORNIA', 'COLORADO',
           'CONNECTICUT', 'DELAWARE', 'FLORIDA', 'GEORGIA', 'HAWAII', 'IDAHO',
           'ILLINOIS', 'INDIANA', 'IOWA', 'KANSAS', 'KENTUCKY', 'LOUISIANA', 'MAINE',
           'MARYLAND', 'MASSACHUSETTS', 'MICHIGAN', 'MINNESOTA', 'MISSISSIPPI', 'MISSOURI',
           'MONTANA', 'NEBRASKA', 'NEVADA', 'NEW HAMPSHIRE', 'NEW JERSEY', 'NEW MEXICO',
           'NEW YORK', 'NORTH CAROLINA', 'NORTH DAKOTA', 'OHIO', 'OKLAHOMA', 'OREGON',
           'PENNSYLVANIA', 'RHODE ISLAND', 'SOUTH CAROLINA', 'SOUTH DAKOTA', 'TENNESSEE',
           'TEXAS', 'UTAH', 'VERMONT', 'VIRGINIA', 'WASHINGTON', 'WEST VIRGINIA', 'WISCONSIN', 'WYOMING', 'DISTRICT OF COLUMBIA', 'AMERICAN SAMOA']


def main():
    # url of search results webpage
    startUrl = 'https://data.bondbuyer.com/salesresults/GetDetails/'
    # the local file path for the results(.tsv and .xls), you need to change it to your local path, like: '/Users/user/Documents/raw_data/linkfilepath/'
    resultPath = 'you need to change'
    # the local file path for the all rew data(.txt), you need to change it to your local path, like: '/Users/user/Documents/raw_data/datafilepath/'
    textPath = 'you need to change'
    # output file path and name for the tsv file of search results webpage
    infoPageName = resultPath + 'infoPage.tsv'
    # start date for the search 
    startDate = '01/01/2000'
    # end date for the search 
    endDate = '11/16/2020'

    get_results_page_to_local(startDate, endDate, infoPageName,textPath)
    get_webpage_to_local(startUrl, textPath, infoPageName)
    
    # file path and name of the fianl results
    outputFile = resultPath + 'final_bondsbuyer.xlsx'
    xlsFile = openpyxl.Workbook()
    fileList = os.listdir(textPath)
    sheet1 = xlsFile.create_sheet(index=0)
    header = ['ID', 'Headline', 'Summary link', 'Publish Date',
              'State', 'Issuer', 'Date', 'Principal', 'Description',
              'Dated', 'Due', 'First Coupon', 'Callable', 'Winning', 'Purchased',
              'Form', 'L.O.', 'F.A.', 'Other Bidder',  'Other Manager']

    for i in range(1, len(header)+1):
        sheet1.cell(1, i).value = header[i-1]

    with open(infoPageName, 'r') as iP:
        i = 2
        par = tqdm.tqdm()
        for line in iP:
            par.update(1)
            tokens = line.split('\t')
            results = []
            if tokens[0].startswith('ID') is False and tokens[1] is not '':
                titleName = tokens[1]+'.txt'
                fileName = textPath + titleName
                with open(fileName, 'r') as f:
                    page = f.read()
                    results = get_pattern(page)
                    for col in range(1, len(tokens)+1):
                        sheet1.cell(i, col).value = tokens[col-1]
                    for res in results:
                        col = len(tokens)+1
                        for v in res:
                            sheet1.cell(i, col).value = res[v]
                            col += 1
                        i += 1
    xlsFile.save(outputFile)


if __name__ == "__main__":
    main()
