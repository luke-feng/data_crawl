# coding: utf-8
from newsplease import NewsPlease
import os
import io
import sys
import json
import tqdm
import pandas as pd
from newsplease import SimpleCrawler
import socket
import copy
import threading
import logging
import datetime
import requests
import urllib3
import openpyxl

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf8')
source_path = 'D:/git/data_crawl/raw_data/'
newslink_path = 'D:/git/data_crawl/raw_data/newslink.csv'
resultPath = source_path + 'news.json'
xl = pd.read_csv(newslink_path, index_col=None, header=None)
urls = xl[0].values.tolist()


def fetch_urls(urls, timeout=None):
    """
    Crawls the html content of all given urls in parallel. Returns when all requests are processed.
    :param urls:
    :param timeout: in seconds, if None, the urllib default is used
    :return:
    """
    threads = [threading.Thread(target=SimpleCrawler._fetch_url, args=(url, True, timeout)) for url in urls if 'www.sec.gov' not in url]
    print(len(threads))
    for thread in threads:
        try:
            thread.start()
        except Exception:
            pass
    for thread in threads:
        try:
            thread.join()
        except Exception:
            pass

    results = copy.deepcopy(SimpleCrawler._results)
    SimpleCrawler._results = {}
    return results

def new_content(url):
    content = dict()
    content["url"] = url
    content["date_publish"] = ''
    content["language"] = ''
    content["source_domain"] = ''
    content["maintext"] = ''
    content["title_NP"] = ''
    content["content_length"] = ''
    return content

def from_urls(urls, timeout=None):
    """
    Crawls articles from the urls and extracts relevant information.
    :param urls:
    :param timeout: in seconds, if None, the urllib default is used
    :return: A dict containing given URLs as keys, and extracted information as corresponding values.
    """
    results = {}
    download_date = datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')

    if len(urls) == 0:
        # Nested blocks of code should not be left empty.
        # When a block contains a comment, this block is not considered to be empty
        pass
    elif len(urls) == 1:
        url = urls[0]
        html = fetch_url(url, timeout=timeout)
        results[url] = NewsPlease.from_html(html, url, download_date)
    else:
        results = fetch_urls(urls, timeout=timeout)
        for url in results:
            try:
                results[url] = NewsPlease.from_html(results[url], url, download_date)
            except Exception:
                results[url] = None
                pass
    return results

def get_content_length(content):
        content_length = 0
        clist = content.split()
        content_length = len(clist)
        return content_length

def crawler():
    with open(resultPath, 'w') as outfile:
        start = []
        end = []
        for i in range(0, len(urls), 10):
            start.append(i)
            if i + 10 >= len(urls):
                end.append(len(urls))
            else:
                end.append(i+10)
        par = tqdm.tqdm(ncols=100, total=len(urls))
        for i, s in enumerate(start):
            results = []
            try:
                url_list=urls[start[i]: end[i]]
                results = from_urls(url_list, timeout=5)
            except Exception as exc:
                pass
            for url in url_list:
                content = new_content(url)
                if url in results:
                    if results[url]!= None and 'www.sec.gov' not in url:
                        content["date_publish"] = str(results[url].date_publish)
                        content["language"] = results[url].language
                        content["source_domain"] = results[url].source_domain
                        content["maintext"] = results[url].maintext
                        content["title_NP"] = results[url].title
                        if content["maintext"] !=None:
                            content["content_length"] = get_content_length(content["maintext"])
                json.dump(content, outfile, indent=4, sort_keys=True)
            par.update(10)
        outfile.close()
    print('get webpage finish!')

def get_content_length(content):
        content_length = 0
        clist = content.split()
        content_length = len(clist)
        return content_length

jsonPath = source_path + 'news.json'
excelPath = source_path + 'news.xlsx'

def from_json_to_excel(jsonPath, excelPath):
    xlsFile = openpyxl.Workbook()
    sheet1 = xlsFile.create_sheet(index=0)
    header = ['url', 'language', 'title', 'date_publish' 'maintext', 'content_length', 'source_domain']

    for i in range(1, len(header)+1):
        sheet1.cell(1, i).value = header[i-1]
    with open(jsonPath, 'r')as old:
        url = ''
        language = ''
        title = ''
        date_publish = ''
        maintext = ''
        content_length = ''
        source_domain = ''
        index = 1
        cl = ''
        jl = ''
        par = tqdm.tqdm(ncols=100, total=len(urls))
        for l in old:
            if '}{' in l:
                par.update(1)
                cl += '}'
                jl = json.loads(cl)
                url = jl['url']
                language = jl['language']
                title = jl['title_NP']
                date_publish = jl['date_publish']
                maintext =jl['maintext']
                content_length =jl['content_length']
                source_domain =jl['source_domain']
                index +=1
                line = [url, language, title, date_publish, maintext, content_length, source_domain]
                for i in range(1, len(line)+1):
                    sheet1.cell(index, i).value = line[i-1]
                cl = '{'
            else:
                cl += l
        xlsFile.save(excelPath)

#crawler()
from_json_to_excel(jsonPath, excelPath)