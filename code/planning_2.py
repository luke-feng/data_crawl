# coding: utf-8

import time
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.support import expected_conditions as Expect
from selenium.webdriver.support.ui import Select
from selenium.webdriver.support.wait import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support.ui import WebDriverWait as Wait
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.by import By
from selenium import webdriver
import requests
from bs4 import BeautifulSoup
import re
import tqdm
import random
import sys
import os
import xlwt
import io
import sys
import urllib.request
import openpyxl
from concurrent.futures import ThreadPoolExecutor
import concurrent.futures
'''
chrome_options = Options()
chrome_options.add_argument('--headless')
# Change your chrome path here
chrome_path = 'C:/Program Files/Google/Chrome/Application/chromedriver.exe'
driver = webdriver.Chrome(
    executable_path=chrome_path, options=chrome_options)
qstart = '01-01-2000'
qend = '01-02-2000'
driver.get('https://online-befirst.lbbd.gov.uk/planning/index.html?fa=search')
fromdate = driver.find_element_by_id('received_date_from')
fromdate.send_keys(qstart)
enddate = driver.find_element_by_id('received_date_to')
enddate.send_keys(qend)
driver.find_element_by_css_selector('#form > div.row.push-20-t > div > div > button.btn.btn-success').click()

for i in range(3):
    driver.execute_script('window.scrollTo(0, document.body.scrollHeight)')
    webdriver.ActionChains(driver).key_down(Keys.DOWN).perform()
    time.sleep(1)
trs = driver.find_elements_by_css_selector('#application_results_table > tbody > tr')
tds = trs[0].find_elements_by_css_selector('td')
refNo = tds[0].text
Status= tds[5].text
dataid = tds[6].find_element_by_css_selector('button').get_attribute('data-id')
link = 'https://online-befirst.lbbd.gov.uk/planning/index.html?fa=getApplication&id=' + str(dataid)
address = tds[2].text
print(refNo,Status,dataid, link, address )
t1 = time.time()
driver.get(link)
t2 = time.time()
print(t2-t1)
rpb5s = driver.find_elements_by_class_name('row.pad-bottom-5')
for rpb5 in  rpb5s:
    md5 = rpb5.find_element_by_class_name('col-md-5').text
    md7 = rpb5.find_element_by_class_name('col-md-7').text
    print(md5, md7)
t3 = time.time()
print(t3-t2)
'''
class Task5:
    def __init__(self, resultPath):
        self.finalResult = []
        self.resultPath = resultPath
        chrome_options = Options()
        chrome_options.add_argument('--headless')
        # Change your chrome path here
        chrome_path = 'C:/Program Files/Google/Chrome/Application/chromedriver.exe'
        self.executor = ThreadPoolExecutor(max_workers=10)
        # for main thread
        self.driver = webdriver.Chrome(
            executable_path=chrome_path, options=chrome_options)
        # for workers
        self.driver1 = webdriver.Chrome(
            executable_path=chrome_path, options=chrome_options)
        self.driver2 = webdriver.Chrome(
            executable_path=chrome_path, options=chrome_options)
        self.driver3 = webdriver.Chrome(
            executable_path=chrome_path, options=chrome_options)
        self.driver4 = webdriver.Chrome(
            executable_path=chrome_path, options=chrome_options)
        self.driver5 = webdriver.Chrome(
            executable_path=chrome_path, options=chrome_options)
        self.driver6 = webdriver.Chrome(
            executable_path=chrome_path, options=chrome_options)
        self.driver7 = webdriver.Chrome(
            executable_path=chrome_path, options=chrome_options)
        self.driver8 = webdriver.Chrome(
            executable_path=chrome_path, options=chrome_options)
        self.driver9 = webdriver.Chrome(
            executable_path=chrome_path, options=chrome_options)
        self.driver10 = webdriver.Chrome(
            executable_path=chrome_path, options=chrome_options)
        self.workers = [self.driver1, self.driver2,
                        self.driver3, self.driver4, self.driver5,self.driver6, self.driver7,
                        self.driver8, self.driver9, self.driver10]
    
    def get_search_results_page(self, url, qstart, qend, council):
        self.driver.get(url)
        fromdate = self.driver.find_element_by_id('received_date_from')
        fromdate.send_keys(qstart)
        enddate = self.driver.find_element_by_id('received_date_to')
        enddate.send_keys(qend)
        self.driver.find_element_by_css_selector('#form > div.row.push-20-t > div > div > button.btn.btn-success').click()
        print(council, qstart, qend)
        par1 = tqdm.tqdm(total=100, ncols=100)
        for i in range(100):
            self.driver.execute_script('window.scrollTo(0, document.body.scrollHeight)')
            webdriver.ActionChains(self.driver).key_down(Keys.DOWN).perform()
            time.sleep(1)
            par1.update(1)
        par1.close()
        trs = self.driver.find_elements_by_css_selector('#application_results_table > tbody > tr')
        return trs


    def run(self, url, council, linkhead):
        header = ['Council', 'Ref. No', 'Name', 'Address', 'Received',
                  'Validated', 'Status', 'MetaInfo', 'Link',
                  'Reference', 'Alternative Reference', 'Application Received',
                  'Application Validated', 'Address', 'Proposal', 'Status',
                  'Decision', 'Decision Issued Date', 'Appeal Status', 'Appeal Decision',
                  'Application Type', 'Decision', 'Actual Decision Level',
                  'Expected Decision Level', 'Case Officer', 'Parish', 'Ward',
                  'District Reference', 'Applicant Name', 'Applicant Address',
                  'Environmental Assessment Requested', 'Agent Name',
                  'Agent Company Name', 'Agent Address', 'Agent Phone Number',
                  'Agent Contacts', 'Councillors',
                  'Application Received Date', 'Application Validated Date',
                  'Expiry Date', 'Actual Committee Date', 'Latest Neighbour Consultation Date',
                  'Neighbour Consultation Expiry Date', 'Standard Consultation Date',
                  'Standard Consultation Expiry Date', 'Last Advertised In Press Date',
                  'Latest Advertisement Expiry Date', 'Last Site Notice Posted Date',
                  'Latest Site Notice Expiry Date', 'Statutory Expiry Date',
                  'Agreed Expiry Date', 'Decision Made Date', 'Decision Issued Date',
                  'Permission Expiry Date', 'Decision Printed Date',
                  'Environmental Impact Assessment Received', 'Temporary Permission Expiry Date', 
                  'Internal Target Date', 'Determination Deadline', 'Target Date', 'Proposed Committee Date'
                  ]

        outputFile = self.resultPath
        xlsFile = openpyxl.Workbook()
        sheet1 = xlsFile.create_sheet(index=0)
        for col in range(1, len(header)+1):
            sheet1.cell(1, col).value = header[col-1]
        line = 2
        for year in range(2000, 2021):
            qstart = '01-01-' + str(year)
            qend = '31-12-' + str(year)
            searchResults = self.get_search_results_page(url, qstart, qend, council)
            par = tqdm.tqdm(total=len(searchResults), ncols=100)
            start = []
            end = []
            for i in range(0, len(searchResults), 10):
                start.append(i)
                if i + 10 >= len(searchResults):
                    end.append(len(searchResults))
                else:
                    end.append(i+10)
            for i, s in enumerate(start):
                self.asyn_page(
                    url_list=searchResults[start[i]:end[i]], council=council, linkhead=linkhead)
                par.update(10)
            par.close()

        for i, line in enumerate(self.finalResult):
            for col in range(1, len(line)+1):
                sheet1.cell(i+2, col).value = line[col-1]
        xlsFile.save(outputFile)
        print('total page is {}'.format(len(self.finalResult)))
        print('get webpage finish!')

    def __del__(self):
        self.driver.close()
        self.driver1.close()
        self.driver2.close()
        self.driver3.close()
        self.driver4.close()
        self.driver5.close()
        self.driver6.close()
        self.driver7.close()
        self.driver8.close()
        self.driver9.close()
        self.driver10.close()
        print('>>>>[Well Done]')

    def get_information(self, browser, result, council, linkhead):
        tds = result.find_elements_by_css_selector('td')
        refNo = tds[0].text
        Received= ''
        Validated= ''
        Status= tds[5].text
        name = ''
        address = tds[2].text
        metaInfo = ''
        details = []
        dataid = tds[6].find_element_by_css_selector('button').get_attribute('data-id')
        link = linkhead + str(dataid)
        details = self.get_details(browser, link)

        res = [council,refNo, name, address, Received,
               Validated, Status, metaInfo, link] + details
        return res

    def get_details(self, browser, url):
        details = []
        browser.get(url)
        Reference = ''
        Alternative_Reference = ''
        Application_Received = ''
        Application_Validated = ''
        Address = ''
        Proposal = ''
        Status = ''
        Decision = ''
        Decision_Issued_Date = ''
        Appeal_Status = ''
        Appeal_Decision = ''
        Application_Type = ''
        Decision = ''
        Actual_Decision_Level = ''
        Expected_Decision_Level = ''
        Case_Officer = ''
        Parish = ''
        Ward = ''
        District_Reference = ''
        Applicant_Name = ''
        Applicant_Address = ''
        Environmental_Assessment_Requested = ''
        Agent_Name = ''
        Agent_Company_Name = ''
        Agent_Address = ''
        Agent_Phone_Number = ''
        agents = ''
        councillors = ''
        Application_Received_Date = ''
        Application_Validated_Date = ''
        Expiry_Date = ''
        Actual_Committee_Date = ''
        Latest_Neighbour_Consultation_Date = ''
        Neighbour_Consultation_Expiry_Date = ''
        Standard_Consultation_Date = ''
        Standard_Consultation_Expiry_Date = ''
        Last_Advertised_In_Press_Date = ''
        Latest_Advertisement_Expiry_Date = ''
        Last_Site_Notice_Posted_Date = ''
        Latest_Site_Notice_Expiry_Date = ''
        Statutory_Expiry_Date = ''
        Agreed_Expiry_Date = ''
        Decision_Made_Date = ''
        Decision_Issued_Date = ''
        Permission_Expiry_Date = ''
        Decision_Printed_Date = ''
        Environmental_Impact_Assessment_Received = ''
        Temporary_Permission_Expiry_Date = ''
        Internal_Target_Date = ''
        Determination_Deadline = ''
        Target_Date = ''
        Proposed_Committee_Date =''

        rpb5s = browser.find_elements_by_class_name('row.pad-bottom-5')
        for rpb5 in  rpb5s:
            key = rpb5.find_element_by_class_name('col-md-5').text
            value = rpb5.find_element_by_class_name('col-md-7').text
            if key == 'Application Lookup:':
                Reference = value
            if key == 'Application Reference Number:':
                Reference = value
            elif key == 'Application Type:':
                Application_Type = value
            elif key == 'Proposal:':
                Proposal = value
            elif key == 'Decision:':
                Decision = value
            elif key == 'Agent:':
                Agent_Name = value
            elif key == 'Applicant:':
                Applicant_Name = value
            elif key == 'Applicant:':
                Applicant_Name = value
            elif key == 'Location:':
                Address = value
            elif key == 'Ward:':
                Ward = value
            elif key == 'Officer:':
                Case_Officer = value
            elif key == 'Received Date:':
                Application_Received_Date = value
            elif key == 'Valid Date:':
                Application_Validated_Date = value
            elif key == 'Proposed Committee Date:':
                Proposed_Committee_Date = value
            elif key == 'Decision Issued Date:':
                Decision_Issued_Date = value
            elif key == 'Application Status:':
                Status = value
               
        details = [Reference, Alternative_Reference, Application_Received, Application_Validated,
                        Address, Proposal, Status, Decision, Decision_Issued_Date, Appeal_Status, Appeal_Decision,
                        Application_Type, Decision, Actual_Decision_Level, Expected_Decision_Level, Case_Officer, Parish, Ward, District_Reference,
                        Applicant_Name, Applicant_Address, Environmental_Assessment_Requested, Agent_Name, Agent_Company_Name, Agent_Address, Agent_Phone_Number,
                       agents, councillors, Application_Received_Date, Application_Validated_Date, Expiry_Date, Actual_Committee_Date,
                     Latest_Neighbour_Consultation_Date, Neighbour_Consultation_Expiry_Date, Standard_Consultation_Date,
                     Standard_Consultation_Expiry_Date, Last_Advertised_In_Press_Date, Latest_Advertisement_Expiry_Date,
                     Last_Site_Notice_Posted_Date, Latest_Site_Notice_Expiry_Date, Statutory_Expiry_Date,
                     Agreed_Expiry_Date, Decision_Made_Date, Decision_Issued_Date, Permission_Expiry_Date,
                     Decision_Printed_Date, Environmental_Impact_Assessment_Received, Temporary_Permission_Expiry_Date, 
                     Internal_Target_Date, Determination_Deadline, Target_Date, Proposed_Committee_Date]
        return details


    def asyn_page(self, url_list, council, linkhead):
        future_to_url  = dict()
        for i, url in enumerate(url_list):
            t = self.executor.submit(self.get_information,
                                 browser=self.workers[i], result=url_list[i], council = council, linkhead=linkhead)
            future_to_url[t] = url               
        for future in concurrent.futures.as_completed(future_to_url):
            url = future_to_url[future]
            try:
                data = future.result()
                self.finalResult.append(data)
            except Exception as exc:
                print('%r generated an exception: %s' % (url, exc))


if __name__ == '__main__':
    # chage the data path here
    datapath = 'D:/git/data_crawl/raw_data/gazette/'
    urllist = [
        'https://online-befirst.lbbd.gov.uk/planning/index.html?fa=search',
        'https://planningapps.hackney.gov.uk/planning/index.html?fa=search',
        'https://builtenvironment.walthamforest.gov.uk/planning/index.html?fa=search'
    ]
    linkhead = ['https://online-befirst.lbbd.gov.uk/planning/index.html?fa=getApplication&id=',
    'https://planningapps.hackney.gov.uk/planning/index.html?fa=getApplication&id=',
    'https://builtenvironment.walthamforest.gov.uk/planning/index.html?fa=getApplication&id='
    ]
    councillist = [
        'London Borough of Barking and Dagenham',
        'London Borough of Hackney',
        'London Borough of Waltham Forest'

    ]
    for i in range(0, len(urllist)):
        try: 
            infoPageName = datapath + councillist[i] + '.xlsx'
            url = urllist[i]
            task = Task5(infoPageName)
            task.run(url,councillist[i], linkhead[i])
        except Exception as exc:
            print(exc)
