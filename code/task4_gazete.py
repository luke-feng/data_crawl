# coding: utf-8

import time
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.support import expected_conditions as Expect
from selenium.webdriver.support.ui import Select
from selenium.webdriver.support.wait import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support.ui import WebDriverWait as Wait
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.by import By
from selenium import webdriver
import requests
from bs4 import BeautifulSoup
import re
import tqdm
import random
import sys
import os
import xlwt
import io
import sys
import urllib.request
import openpyxl
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf8')


def get_all_article(upper, resultPath):
    """
    get web source code from the results webpage
    :param url: the results webpage
    :return page: string data type, page source code
    """

    chrome_options = Options()
    chrome_options.add_argument('--headless')
    browser = webdriver.Chrome(
        executable_path="C:/Program Files/Google/Chrome/Application/chromedriver.exe", options=chrome_options)
    url = 'https://www.thegazette.co.uk/insolvency/notice?text=&insolvency_corporate=G205010000&insolvency_personal\
        =G206030000&location-postcode-1=&location-distance-1=1&location-local-authority-1=&numberOfLocationSearches=1&\
        start-publish-date=&end-publish-date=&edition=&london-issue=&edinburgh-issue=&belfast-issue=&sort-by=&results-page-size=10&results-page='

    outputFile = resultPath
    xlsFile = openpyxl.Workbook()
    sheet1 = xlsFile.create_sheet(index=0)
    header = ['Item ID', 'Publication Date', 'Title', 'Summary',
              'Notice Type', 'Link', 'Notice Category', 'Notice Type', 'Earliest Publish Date',
              'Publication Date', 'Edition', 'Notice ID', 'Company Number', 'Company Link',
              'Notice Code', 'More', 'Timeline Title', 'TimeLine', 'Detail Title', 'Detail Content']
    for col in range(1, len(header)+1):
        sheet1.cell(1, col).value = header[col-1]
    line = 2
    for i in range(1, upper+1):
        url = url + str(i)
        browser.get(url)
        article = browser.find_elements_by_css_selector("body>.wrapper>.wrapperContent>\
                .main-group.no-hero>#main_content>.services-content>\
                #searchform>.main-pane>section>#search-results>.content>article")
        # get web source code
        if len(article) > 1:
            for tr in article:
                trId = tr.get_attribute('id')
                trPd = tr.find_element_by_css_selector(
                    '.feed-item>.metadata.publication-date>dd>time').text
                trTi = tr.find_element_by_css_selector(
                    '.feed-item>header>h3>a').text
                trSu = tr.find_element_by_css_selector(
                    '.feed-item>.content>div').text
                trSu = trSu.replace('\n', ' ')
                trNt = tr.find_element_by_css_selector(
                    '.feed-item>.metadata.notice-type>dd').text
                trL = tr.find_element_by_css_selector(
                    '.feed-item>a').get_attribute('href')
                print(trL)
                notice_information = get_information(trL)
                result = [trId, trPd, trTi, trSu,
                          trNt, trL] + notice_information
                for col in range(1, len(header)+1):
                    sheet1.cell(line, col).value = result[col-1]
                line += 1

    xlsFile.save(outputFile)
    print('total page is {}'.format(str(line-1)))
    browser.close()
    print('get webpage finish!')


datapath = 'D:/git/data_crawl/raw_data/gazette/'
textPath = datapath+'webpage/'
infoPageName = datapath + 'result.xlsx'


def get_information(url):
    chrome_options = Options()
    chrome_options.add_argument('--headless')
    browser = webdriver.Chrome(
        executable_path="C:/Program Files/Google/Chrome/Application/chromedriver.exe", options=chrome_options)
    browser.get(url)
    information = browser.find_element_by_css_selector(
        'body>.wrapper>.wrapperContent>.main-group.no-hero>#main_content>.services-content')
    results = []
    nd = get_notice_details(
        information)
    nt = get_notice_timeline(
        information)
    nc = get_notice(information)
    results = nd + nt + nc
    browser.close()
    return results


def get_notice_details(information):
    notice_details = information.find_element_by_css_selector(
        '.notice-wrapper>.related-pane>section>.notice-summary')
    category = ''
    notice_type = ''
    earliest_publish_date = ''
    publication_date = ''
    edition = ''
    noticeID = ''
    company_number = ''
    company_link = ''
    notice_code = ''
    about = ''
    results = []
    '''category = notice_details.find_element_by_css_selector(
        '.metadata>dd.category').text
    notice_type = notice_details.find_element_by_css_selector(
        '.metadata>dd.notice-type').text'''
    dds = notice_details.find_elements_by_css_selector(
        '.metadata>dd')
    if len(dds) == 7:
        category = dds[0].text
        notice_type = dds[1].text
        earliest_publish_date = dds[2].text
        edition = dds[3].text
        noticeID = dds[4].text
        company_number = dds[5].text
        company_link = dds[5].find_element_by_css_selector(
            'a').get_attribute('href')
        notice_code = dds[6].text
        about_raw = notice_details.find_elements_by_css_selector('.more')
        if len(about_raw) > 0:
            about = about_raw[0].text+' ' + about_raw[0].get_attribute('href')
    elif len(dds) == 6:
        category = dds[0].text
        notice_type = dds[1].text
        publication_date = dds[2].text
        edition = dds[3].text
        noticeID = dds[4].text
        notice_code = dds[5].text
    else:
        print('pattern error!!!')

    results = [category, notice_type, earliest_publish_date, publication_date,
               edition, noticeID, company_number, company_link, notice_code, about]
    return results


def get_notice_timeline(information):
    aside = information.find_elements_by_css_selector(
        '.notice-wrapper>.main-pane.no-focus>aside')
    timeline_title = ''
    timeline = ''
    results = []
    if len(aside) > 0:
        notice_timeline = aside[0].find_element_by_css_selector(
            '.notice-timeline')
        timeline_title = notice_timeline.find_element_by_css_selector(
            '.title.timeline-title').text
        scroll_pane = notice_timeline.find_element_by_css_selector(
            '.timeline.scroll-pane>.jspContainer>.jspPane>ol')
        lis = scroll_pane.find_elements_by_css_selector('li')
        for li in lis:
            cl = li.get_attribute('class')
            if 'current' in cl:
                item_link = ''
                item_id = ''
                item_title = li.find_element_by_css_selector(
                    '.item-title').text
                item_datatime = li.find_element_by_css_selector('time').text
            else:
                item_link = li.find_element_by_css_selector(
                    'a').get_attribute('href')
                item_id = item_link.split('/')[4]
                item_title = li.find_element_by_css_selector(
                    'a>.item-title').text
                item_datatime = li.find_element_by_css_selector('time').text

            line = 'Item title: {}, Item ID: {}, Link: {}, Datetime: {}'.format(
                item_title, item_id, item_link, item_datatime)
            results += line+'\n'
        #print('{} {} {} {}'.format(timeline_title, meeting, resolution, appointment))
        results = [timeline_title,timeline]
        return results

    else:
        results = [timeline_title, timeline]
        return results


def get_notice(information):
    full_notice = information.find_element_by_css_selector(
        '.notice-wrapper>.main-pane.no-focus>article>div')
    title = full_notice.find_element_by_css_selector('header>.title').text
    content = ''
    results = []
    divs = full_notice.find_elements_by_css_selector('.content>div')
    for div in divs:
        content += div.text + '\n'
    results = [title, content]
    return results


def save_page(path, ID, page):
    filename = path + ID + '.txt'
    with open(filename, 'w') as f:
        f.write('{}'.format(page.encode('utf-8', 'ignore')))
        f.close()


#information = get_information('https://www.thegazette.co.uk/notice/3736042')
get_all_article(1, infoPageName)
