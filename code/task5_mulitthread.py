# coding: utf-8

import time
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.support import expected_conditions as Expect
from selenium.webdriver.support.ui import Select
from selenium.webdriver.support.wait import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support.ui import WebDriverWait as Wait
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.by import By
from selenium import webdriver
import requests
from bs4 import BeautifulSoup
import re
import tqdm
import random
import sys
import os
import xlwt
import io
import sys
import urllib.request
import openpyxl
from concurrent.futures import ThreadPoolExecutor
import concurrent.futures
finalResult = []


class Task5:
    def __init__(self, resultPath):
        self.resultPath = resultPath
        chrome_options = Options()
        chrome_options.add_argument('--headless')

        self.executor = ThreadPoolExecutor(max_workers=10)
        # for main thread
        self.driver = webdriver.Chrome(
            executable_path='C:/Program Files/Google/Chrome/Application/chromedriver.exe', options=chrome_options)
        # for workers
        self.driver1 = webdriver.Chrome(
            executable_path='C:/Program Files/Google/Chrome/Application/chromedriver.exe', options=chrome_options)
        self.driver2 = webdriver.Chrome(
            executable_path='C:/Program Files/Google/Chrome/Application/chromedriver.exe', options=chrome_options)
        self.driver3 = webdriver.Chrome(
            executable_path='C:/Program Files/Google/Chrome/Application/chromedriver.exe', options=chrome_options)
        self.driver4 = webdriver.Chrome(
            executable_path='C:/Program Files/Google/Chrome/Application/chromedriver.exe', options=chrome_options)
        self.driver5 = webdriver.Chrome(
            executable_path='C:/Program Files/Google/Chrome/Application/chromedriver.exe', options=chrome_options)
        self.driver6 = webdriver.Chrome(
            executable_path='C:/Program Files/Google/Chrome/Application/chromedriver.exe', options=chrome_options)
        self.driver7 = webdriver.Chrome(
            executable_path='C:/Program Files/Google/Chrome/Application/chromedriver.exe', options=chrome_options)
        self.driver8 = webdriver.Chrome(
            executable_path='C:/Program Files/Google/Chrome/Application/chromedriver.exe', options=chrome_options)
        self.driver9 = webdriver.Chrome(
            executable_path='C:/Program Files/Google/Chrome/Application/chromedriver.exe', options=chrome_options)
        self.driver10 = webdriver.Chrome(
            executable_path='C:/Program Files/Google/Chrome/Application/chromedriver.exe', options=chrome_options)
        self.workers = [self.driver1, self.driver2,
                        self.driver3, self.driver4, self.driver5,self.driver6, self.driver7,
                        self.driver8, self.driver9, self.driver10]

    def run(self):

        url = 'https://www1.somersetwestandtaunton.gov.uk/online-applications/search.do?action=advanced'
        startday = ['01/01', '01/02', '01/03', '01/04', '01/05', '01/06',
                    '01/07', '01/08', '01/09', '01/10', '01/11', '01/12']
        endday = ['01/02', '01/03', '01/04', '01/05', '01/06', '01/07',
                  '01/08', '01/09', '01/10', '01/11', '01/12', '31/12']

        header = ['Ref. No', 'Name', 'Address', 'Received',
                  'Validated', 'Status', 'MetaInfo', 'Link',
                  'Reference', 'Alternative Reference', 'Application Received',
                  'Application Validated', 'Address', 'Proposal', 'Status',
                  'Decision', 'Decision Issued Date', 'Appeal Status', 'Appeal Decision',
                  'Application Type', 'Decision', 'Actual Decision Level',
                  'Expected Decision Level', 'Case Officer', 'Parish', 'Ward',
                  'District Reference', 'Applicant Name', 'Applicant Address',
                  'Environmental Assessment Requested', 'Agent Name',
                  'Agent Company Name', 'Agent Address', 'Agent Phone Number',
                  'Agent Contacts', 'Councillors',
                  'Application Received Date', 'Application Validated Date',
                  'Expiry Date', 'Actual Committee Date', 'Latest Neighbour Consultation Date',
                  'Neighbour Consultation Expiry Date', 'Standard Consultation Date',
                  'Standard Consultation Expiry Date', 'Last Advertised In Press Date',
                  'Latest Advertisement Expiry Date', 'Last Site Notice Posted Date',
                  'Latest Site Notice Expiry Date', 'Statutory Expiry Date',
                  'Agreed Expiry Date', 'Decision Made Date', 'Decision Issued Date',
                  'Permission Expiry Date', 'Decision Printed Date',
                  'Environmental Impact Assessment Received', 'Temporary Permission Expiry Date'
                  ]

        outputFile = self.resultPath
        xlsFile = openpyxl.Workbook()
        sheet1 = xlsFile.create_sheet(index=0)
        for col in range(1, len(header)+1):
            sheet1.cell(1, col).value = header[col-1]
        line = 2

        for year in range(2000, 2020):
            for mon in range(0, 12):
                qstart = startday[mon]+'/'+str(year)
                qend = endday[mon]+'/'+str(year)
                print(qstart, qend)
                self.driver.get(url)
                Wait(self.driver, 600).until(
                    Expect.presence_of_element_located(
                        (By.CSS_SELECTOR, 'body>#idox>#pa>.container>.content>.tabcontainer>#advancedSearchForm>#dates>fieldset>.row>.col-dateFrom'))
                )
                fromdate = self.driver.find_element_by_id(
                    'applicationReceivedStart')
                fromdate.send_keys(qstart)
                enddate = self.driver.find_element_by_id(
                    'applicationReceivedEnd')
                enddate.send_keys(qend)
                self.driver.find_element_by_css_selector(
                    'body>#idox>#pa>.container>.content>.tabcontainer>#advancedSearchForm>.buttons>.button.primary').click()
                self.driver.get(self.driver.current_url)
                Wait(self.driver, 600).until(
                    Expect.presence_of_element_located(
                        (By.ID, 'resultsPerPage'))
                )
                resultsPerPage = self.driver.find_element_by_id(
                    'resultsPerPage')
                s1 = Select(resultsPerPage)
                s1.select_by_value('100')
                self.driver.find_element_by_css_selector(
                    'body>#idox>#pa>.container>.content>#searchfilters>#searchResults>.button.primary').click()
                # browser.get(browser.current_url)
                searchResults = self.driver.find_elements_by_class_name(
                    'searchresult')

                while True:
                    next = self.driver.find_elements_by_class_name('next')
                    if len(next) > 0:
                        par = tqdm.tqdm(total=len(searchResults), ncols=100)
                        start = []
                        end = []
                        for i in range(0, len(searchResults), 10):
                            start.append(i)
                            if i + 10 >= len(searchResults):
                                end.append(len(searchResults))
                            else:
                                end.append(i+10)
                        for i, s in enumerate(start):
                            self.asyn_page(
                                url_list=searchResult[start[i], end[i]])
                            par.update(10)
                        par.close()
                        next_url = next[0].get_attribute('href')
                        self.driver.get(next_url)
                        searchResults = self.driver.find_elements_by_class_name(
                            'searchresult')
                    else:
                        par = tqdm.tqdm(total=len(searchResults), ncols=100)
                        start = []
                        end = []
                        for i in range(0, len(searchResults), 10):
                            start.append(i)
                            if i + 10 >= len(searchResults):
                                end.append(len(searchResults))
                            else:
                                end.append(i+10)
                        for i, s in enumerate(start):
                            self.asyn_page(
                                url_list=searchResults[start[i]:end[i]])
                            par.update(10)
                        par.close()
                        break
        for i, line in enumerate(finalResult):
            for col in range(1, len(line)+1):
                sheet1.cell(i+2, col).value = line[col-1]
        xlsFile.save(outputFile)
        print('total page is {}'.format(len(finalResult)))
        print('get webpage finish!')

    def __del__(self):
        self.driver.close()
        self.driver1.close()
        self.driver2.close()
        self.driver3.close()
        self.driver4.close()
        self.driver5.close()
        self.driver6.close()
        self.driver7.close()
        self.driver8.close()
        self.driver9.close()
        self.driver10.close()
        print('>>>>[Well Done]')

    def get_information(self, browser, result):
        link = result.find_element_by_css_selector(
            'a').get_attribute('href')
        name = result.find_element_by_css_selector(
            'a').text
        address = result.find_element_by_css_selector(
            '.address').text
        metaInfo = result.find_element_by_css_selector(
            '.metaInfo').text
        ml = metaInfo.split('|')
        refNo = ml[0].split(':')[1]
        Received = ml[1].split(':')[1]
        Validated = ml[2].split(':')[1]
        Status = ml[3].split(':')[1]
        summary = []
        details = []
        contacts = []
        dates = []
        results = []
        summarylink = link
        detailslink = link.replace('activeTab=summary', 'activeTab=details')
        contactslink = link.replace('activeTab=summary', 'activeTab=contacts')
        dateslink = link.replace('activeTab=summary', 'activeTab=dates')
        summary = self.get_summary(browser, summarylink)
        details = self.get_details(browser, detailslink)
        contacts = self.get_contacts(browser, contactslink)
        dates = self.get_dates(browser, dateslink)
        res = [refNo, name, address, Received,
               Validated, Status, metaInfo, link] + summary + details + contacts + dates
        return res

    def get_summary(self, browser, url):
        summary = []
        browser.get(url)
        Reference = ''
        Alternative_Reference = ''
        Application_Received = ''
        Application_Validated = ''
        Address = ''
        Proposal = ''
        Status = ''
        Decision = ''
        Decision_Issued_Date = ''
        Appeal_Status = ''
        Appeal_Decision = ''

        simpleDetailsTable = browser.find_elements_by_id('simpleDetailsTable')
        if len(simpleDetailsTable) > 0:
            trs = simpleDetailsTable[0].find_elements_by_css_selector(
                'tbody>tr')
            for tr in trs:
                content = tr.find_element_by_css_selector('td').text
                key = tr.find_element_by_css_selector('th').text
                if key == 'Reference':
                    Reference = content
                elif key == 'Alternative Reference':
                    Alternative_Reference = content
                elif key == 'Application Received':
                    Application_Received = content
                elif key == 'Application Validated':
                    Application_Validated = content
                elif key == 'Address':
                    Address = content
                elif key == 'Proposal':
                    Proposal = content
                elif key == 'Status':
                    Status = content
                elif key == 'Decision':
                    Decision = content
                elif key == 'Decision Issued Date':
                    Decision_Issued_Date = content
                elif key == 'Appeal Status':
                    Appeal_Status = content
                elif key == 'Appeal Decision':
                    Appeal_Decision = content
        summary = [Reference, Alternative_Reference, Application_Received, Application_Validated,
                   Address, Proposal, Status, Decision, Decision_Issued_Date, Appeal_Status, Appeal_Decision]
        return summary

    def get_details(self, browser, url):
        details = []
        browser.get(url)
        Application_Type = ''
        Decision = ''
        Actual_Decision_Level = ''
        Expected_Decision_Level = ''
        Case_Officer = ''
        Parish = ''
        Ward = ''
        District_Reference = ''
        Applicant_Name = ''
        Applicant_Address = ''
        Environmental_Assessment_Requested = ''
        Agent_Name = ''
        Agent_Company_Name = ''
        Agent_Address = ''
        Agent_Phone_Number = ''
        simpleDetailsTable = browser.find_elements_by_id('applicationDetails')
        if len(simpleDetailsTable) > 0:
            trs = simpleDetailsTable[0].find_elements_by_css_selector(
                'tbody>tr')
            for tr in trs:
                content = tr.find_element_by_css_selector('td').text
                key = tr.find_element_by_css_selector('th').text
                if key == 'Application Type':
                    Application_Type = content
                elif key == 'Decision':
                    Decision = content
                elif key == 'Actual Decision Level':
                    Actual_Decision_Level = content
                elif key == 'Expected Decision Level':
                    Expected_Decision_Level = content
                elif key == 'Case Officer':
                    Case_Officer = content
                elif key == 'Parish':
                    Parish = content
                elif key == 'Ward':
                    Ward = content
                elif key == 'District Reference':
                    District_Reference = content
                elif key == 'Applicant Name':
                    Applicant_Name = content
                elif key == 'Applicant Address':
                    Applicant_Address = content
                elif key == 'Environmental Assessment Requested':
                    Environmental_Assessment_Requested = content
                elif key == 'Agent Name':
                    Agent_Name = content
                elif key == 'Agent Company Name':
                    Agent_Company_Name = content
                elif key == 'Agent Address':
                    Agent_Address = content
                elif key == 'Agent Phone Number':
                    Agent_Phone_Number = content
            details = [Application_Type, Decision, Actual_Decision_Level, Expected_Decision_Level, Case_Officer, Parish, Ward, District_Reference,
                       Applicant_Name, Applicant_Address, Environmental_Assessment_Requested, Agent_Name, Agent_Company_Name, Agent_Address, Agent_Phone_Number]
        return details

    def get_contacts(self, browser, url):
        contacts = []
        browser.get(url)
        agents = ''
        councillors = ''
        agentsTable = browser.find_elements_by_class_name(
            'tabcontainer>.agents')
        councillorsTable = browser.find_elements_by_class_name(
            'tabcontainer>.councillors')
        if len(agentsTable) > 0:
            for agt in agentsTable:
                agents += agt.text + '\n'
        if len(councillorsTable) > 0:
            for cot in councillorsTable:
                councillors += cot.text + '\n'
        contacts = [agents, councillors]
        return contacts

    def get_dates(self, browser, url):
        dates = []
        browser.get(url)
        Application_Received_Date = ''
        Application_Validated_Date = ''
        Expiry_Date = ''
        Actual_Committee_Date = ''
        Latest_Neighbour_Consultation_Date = ''
        Neighbour_Consultation_Expiry_Date = ''
        Standard_Consultation_Date = ''
        Standard_Consultation_Expiry_Date = ''
        Last_Advertised_In_Press_Date = ''
        Latest_Advertisement_Expiry_Date = ''
        Last_Site_Notice_Posted_Date = ''
        Latest_Site_Notice_Expiry_Date = ''
        Statutory_Expiry_Date = ''
        Agreed_Expiry_Date = ''
        Decision_Made_Date = ''
        Decision_Issued_Date = ''
        Permission_Expiry_Date = ''
        Decision_Printed_Date = ''
        Environmental_Impact_Assessment_Received = ''
        Temporary_Permission_Expiry_Date = ''

        simpleDetailsTable = browser.find_elements_by_id('simpleDetailsTable')
        if len(simpleDetailsTable) > 0:
            trs = simpleDetailsTable[0].find_elements_by_css_selector(
                'tbody>tr')
            for tr in trs:
                content = tr.find_element_by_css_selector('td').text
                key = tr.find_element_by_css_selector('th').text
                if key == 'Application Received Date':
                    Application_Received_Date = content
                elif key == 'Application Validated Date':
                    Application_Validated_Date = content
                elif key == 'Expiry Date':
                    Expiry_Date = content
                elif key == 'Actual Committee Date':
                    Actual_Committee_Date = content
                elif key == 'Latest Neighbour Consultation Date':
                    Latest_Neighbour_Consultation_Date = content
                elif key == 'Neighbour Consultation Expiry Date':
                    Neighbour_Consultation_Expiry_Date = content
                elif key == 'Standard Consultation Date':
                    Standard_Consultation_Date = content
                elif key == 'Standard Consultation Expiry Date':
                    Standard_Consultation_Expiry_Date = content
                elif key == 'Last Advertised In Press Date':
                    Last_Advertised_In_Press_Date = content
                elif key == 'Latest Advertisement Expiry Date':
                    Latest_Advertisement_Expiry_Date = content
                elif key == 'Last Site Notice Posted Date':
                    Last_Site_Notice_Posted_Date = content
                elif key == 'Latest Site Notice Expiry Date':
                    Latest_Site_Notice_Expiry_Date = content
                elif key == 'Statutory Expiry Date':
                    Statutory_Expiry_Date = content
                elif key == 'Agreed Expiry Date':
                    Agreed_Expiry_Date = content
                elif key == 'Decision Made Date':
                    Decision_Made_Date = content
                elif key == 'Decision Issued Date':
                    Decision_Issued_Date = content
                elif key == 'Permission Expiry Date':
                    Permission_Expiry_Date = content
                elif key == 'Decision Printed Date':
                    Decision_Printed_Date = content
                elif key == 'Environmental Impact Assessment Received':
                    Environmental_Impact_Assessment_Received = content
                elif key == 'Temporary Permission Expiry Date':
                    Temporary_Permission_Expiry_Date = content
            dates = [Application_Received_Date, Application_Validated_Date, Expiry_Date, Actual_Committee_Date,
                     Latest_Neighbour_Consultation_Date, Neighbour_Consultation_Expiry_Date, Standard_Consultation_Date,
                     Standard_Consultation_Expiry_Date, Last_Advertised_In_Press_Date, Latest_Advertisement_Expiry_Date,
                     Last_Site_Notice_Posted_Date, Latest_Site_Notice_Expiry_Date, Statutory_Expiry_Date,
                     Agreed_Expiry_Date, Decision_Made_Date, Decision_Issued_Date, Permission_Expiry_Date,
                     Decision_Printed_Date, Environmental_Impact_Assessment_Received, Temporary_Permission_Expiry_Date]
        return dates

    def asyn_page(self, url_list):
        future_to_url  = dict()
        for i, url in enumerate(url_list):
            t = self.executor.submit(self.get_information,
                                 browser=self.workers[i], result=url_list[i])
            future_to_url[t] = url               
        for future in concurrent.futures.as_completed(future_to_url):
            url = future_to_url[future]
            try:
                data = future.result()
                finalResult.append(data)
            except Exception as exc:
                print('%r generated an exception: %s' % (url, exc))


if __name__ == '__main__':
    datapath = 'D:/git/data_crawl/raw_data/gazette/'
    infoPageName = datapath + 'result1.xlsx'
    task = Task5(infoPageName)
    task.run()
