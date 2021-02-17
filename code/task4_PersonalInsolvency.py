# coding: utf-8

import time
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.support import expected_conditions as Expect
from selenium.webdriver.support.ui import Select
from selenium.webdriver.support.wait import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support.ui import WebDriverWait as Wait
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.by import By
from selenium import webdriver
import requests
from bs4 import BeautifulSoup
import re
import tqdm
import random
import sys
import os
import xlwt
import io
import sys
import urllib.request
import openpyxl
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf8')


def get_all_article(upper, resultPath):
    '''
    get web source code from the results webpage
    :param url: the results webpage
    :return page: string data type, page source code
    '''

    chrome_options = Options()
    chrome_options.add_argument('--headless')
    browser = webdriver.Chrome(
        executable_path='C:/Program Files/Google/Chrome/Application/chromedriver.exe', options=chrome_options)
    browser1 = webdriver.Chrome(
        executable_path='C:/Program Files/Google/Chrome/Application/chromedriver.exe', options=chrome_options)

    url = 'https://www.thegazette.co.uk/insolvency/notice?text=&categorycode=-1&insolvency_personal=G206030000&\
        location-postcode-1=&location-distance-1=1&location-local-authority-1=&numberOfLocationSearches=1&\
            start-publish-date=&end-publish-date=&edition=&london-issue=&edinburgh-issue=&belfast-issue=&sort-by=&results-page-size=100&results-page='

    outputFile = resultPath
    xlsFile = openpyxl.Workbook()
    sheet1 = xlsFile.create_sheet(index=0)
    header = ['Item ID', 'Publication Date', 'Title', 'Summary',
              'Notice Type', 'Link', 'Notice Category', 'Notice Type', 'Earliest Publish Date',
              'Publication Date', 'Edition', 'Notice ID', 'Notice Code', 'Issue Number', 'Page Number',
              'More', 'Court Name', 'Case Code', 'Name', 'Also Known As', 'Address', 'Date Of Birth', 'Date Of Appointment',
              'Status', 'Date Of Bankruptcy Order', 'Detail Title', 'Detail Content']

    for col in range(1, len(header)+1):
        sheet1.cell(1, col).value = header[col-1]
    line = 2
    par = tqdm.tqdm()
    for i in range(1, upper+1):
        url = url + str(i)
        browser.get(url)
        Wait(browser, 600).until(
            Expect.presence_of_element_located(
                (By.CSS_SELECTOR, 'body>.wrapper>.wrapperContent>\
                .main-group.no-hero>#main_content>.services-content>\
                #searchform>.main-pane>section>#search-results>.content>article'))
        )
        article = browser.find_elements_by_css_selector('body>.wrapper>.wrapperContent>\
                .main-group.no-hero>#main_content>.services-content>\
                #searchform>.main-pane>section>#search-results>.content>article')
        # get web source code
        if len(article) > 1:
            for tr in article:
                par.update(1)
                trId = tr.get_attribute('id')
                trPd = tr.find_element_by_css_selector(
                    '.feed-item>.metadata.publication-date>dd>time').text
                trTi = tr.find_element_by_css_selector(
                    '.feed-item>header>h3>a').text
                trSu = tr.find_element_by_css_selector(
                    '.feed-item>.content>div').text
                trSu = trSu.replace('\n', ' ')
                trNt = tr.find_element_by_css_selector(
                    '.feed-item>.metadata.notice-type>dd').text
                trL = tr.find_element_by_css_selector(
                    '.feed-item>a').get_attribute('href')
                print(trL)
                notice_information = get_information(browser1, trL)
                result = [trId, trPd, trTi, trSu,
                          trNt, trL] + notice_information
                for col in range(1, len(header)+1):
                    sheet1.cell(line, col).value = result[col-1]
                line += 1

    xlsFile.save(outputFile)
    print('total page is {}'.format(str(line-1)))
    browser.close()
    browser1.close()
    par.close()
    print('get webpage finish!')

def get_information(browser, url):
    browser.get(url)
    Wait(browser, 600).until(
        Expect.presence_of_element_located(
            (By.CSS_SELECTOR, 'body>.wrapper>.wrapperContent>.main-group.no-hero>#main_content>.services-content'))
    )
    information = browser.find_element_by_css_selector(
        'body>.wrapper>.wrapperContent>.main-group.no-hero>#main_content>.services-content')
    results = []
    nd = get_notice_details(
        information)
    nc = get_notice_whole(information)
    results = nd + nc
    return results


def get_notice_details(information):
    notice_details = information.find_element_by_css_selector(
        '.notice-wrapper>.related-pane>section>.notice-summary')
    category = ''
    notice_type = ''
    earliest_publish_date = ''
    publication_date = ''
    edition = ''
    noticeID = ''
    notice_code = ''
    issue_number = ''
    page_number = ''
    about = ''
    results = []
    dds = notice_details.find_elements_by_css_selector(
        '.metadata>dd')
    dts = notice_details.find_elements_by_css_selector(
        '.metadata>dt')
    if len(dds) != len(dts):
        print('error!!!')
    else:
        for i, dt in enumerate(dts):
            if 'Type:' in dt.text:
                category = dds[i].text
            elif 'Notice type:' in dt.text:
                notice_type = dds[i].text
            elif 'Earliest publish date:' in dt.text:
                earliest_publish_date = dds[i].text
            elif 'Publication date:' in dt.text:
                publication_date = dds[i].text
            elif 'Edition:' in dt.text:
                edition = dds[i].text
            elif 'Notice ID:' in dt.text:
                noticeID = dds[i].text
            elif 'Notice code:' in dt.text:
                notice_code = dds[i].text
            elif 'Issue number:' in dt.text:
                issue_number = dds[i].text
            elif 'Page number:' in dt.text:
                page_number = dds[i].text
        about_raw = notice_details.find_elements_by_css_selector('.more')
        if len(about_raw) > 0:
            about = about_raw[0].text+' ' + about_raw[0].get_attribute('href')
    results = [category, notice_type, earliest_publish_date, publication_date,
               edition, noticeID, notice_code, issue_number, page_number, about]
    return results


def get_notice_whole(information):
    full_notice = information.find_element_by_css_selector(
        '.notice-wrapper>.main-pane.no-focus>article>div')
    title = full_notice.find_element_by_css_selector('header>.title').text
    content = ''
    results = []
    divs = full_notice.find_elements_by_css_selector('.content>div')
    for div in divs:
        content += div.text + '\n'
    keyvalue = get_notice_keyvalue(divs)
    results = keyvalue + [title, content]
    return results


def get_notice_keyvalue(divs):
    sp = dict()
    for div in divs:
        spans = div.find_elements_by_tag_name('span')
        for span in spans:
            key = span.get_attribute('property')
            value = span.text
            if key not in sp:
                sp[key] = value
            #print('key:{}, Value:{}'.format(key, value))
        ps = div.find_elements_by_tag_name('p')
        for p in ps:
            key = p.get_attribute('property')
            value = p.text
            if key not in sp:
                sp[key] = value
            #print('key:{}, Value:{}'.format(key, value))
        h3s = div.find_elements_by_tag_name('h3')
        for h3 in h3s:
            key = 'h3 '+h3.get_attribute('data-gazettes')
            value = h3.text
            if key not in sp:
                sp[key] = value
            #print('key:{}, Value:{}'.format(key, value))
        h2s = div.find_elements_by_tag_name('h2')
        for h2 in h2s:
            key = 'h2 '+h2.get_attribute('data-gazettes')
            value = h2.text
            if key not in sp:
                sp[key] = value
            #print('key:{}, Value:{}'.format(key, value))

    courtName = ''
    caseCode = ''
    name = ''
    hasAddress = ''
    hasStatus = ''
    dateOfAppointment = ''
    dateOfBankruptcyOrder = ''
    dateOfBirth = ''
    alsoKnownAs = ''
    firstName = ''
    givenName = ''
    familyName = ''
    caseNumber = ''
    caseYear = ''
    street_address = ''
    extended_address = ''
    locality = ''
    postal_code = ''
    for key in sp:
        if key is None:
            continue
        elif 'courtName' in key:
            courtName = sp[key]
        elif 'caseCode' in key:
            caseCode = sp[key]
        elif 'h3' in key:
            name = sp[key]
        elif 'h2' in key:
            name = sp[key]
        elif 'hasAddress' in key:
            hasAddress = sp[key]
        elif 'hasStatus' in key:
            hasStatus = sp[key]
        elif 'dateOfAppointment' in key:
            dateOfAppointment = sp[key]
        elif 'dateOfBankruptcyOrder' in key:
            dateOfBankruptcyOrder = sp[key]
        elif 'dateOfBirth' in key:
            dateOfBirth = sp[key]
        elif 'alsoKnownAs' in key:
            alsoKnownAs = sp[key]
        elif 'firstName' in key:
            firstName = sp[key]
        elif 'givenName' in key:
            givenName = sp[key]
        elif 'familyName' in key:
            familyName = sp[key]
        elif 'caseNumber' in key:
            caseNumber = sp[key]
        elif 'caseYear' in key:
            caseYear = sp[key]
        elif 'street-address' in key:
            street_address = sp[key]
        elif 'extended-address' in key:
            extended_address = sp[key]
        elif 'locality' in key:
            locality = sp[key]
        elif 'postal-code' in key:
            postal_code = sp[key]

    if caseCode == '' and caseNumber != '':
        caseCode = caseNumber + ' of ' + caseYear
    if name == '' and firstName != '':
        name = firstName + ' ' + givenName + ' ' + familyName
    if hasAddress == '' and street_address != '':
        hasAddress = street_address + ', ' + extended_address + \
            ', ' + locality + ', ' + postal_code
    hasAddress = hasAddress.replace(', ,', ',')
    Appointment = dateOfAppointment.split(':')
    BankruptcyOrder = dateOfBankruptcyOrder.split(':')
    Birth = dateOfBirth.split(':')
    nc = name.split('(')
    if len(Appointment) > 1:
        dateOfAppointment = Appointment[1]
    if len(BankruptcyOrder) > 1:
        dateOfBankruptcyOrder = BankruptcyOrder[1]
    if len(Birth) > 1:
        dateOfBirth = Birth[1]
    if len(nc) > 1:
        name = nc[1]
    result = [courtName, caseCode, name, alsoKnownAs, hasAddress,
              dateOfBirth, dateOfAppointment, hasStatus, dateOfBankruptcyOrder]
    return result


#information = get_information('https://www.thegazette.co.uk/notice/3734984')
datapath = 'D:/git/data_crawl/raw_data/gazette/'
infoPageName = datapath + 'insolvency_personal_demo100.xlsx'
get_all_article(1, infoPageName)
