# coding: utf-8

import time
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.support import expected_conditions as Expect
from selenium.webdriver.support.ui import Select
from selenium.webdriver.support.wait import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support.ui import WebDriverWait as Wait
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.by import By
from selenium import webdriver
import requests
from bs4 import BeautifulSoup
import re
import tqdm
import random
import sys
import os
import xlwt
import io
import sys
import urllib.request
import openpyxl
import pandas as pd
from random import randint

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf8')

# !change the Trademark_ID_List.xlsx file path here
id_path = 'D:/git/data_crawl/raw_data/Trademark_ID_List.xlsx'
# !change the chromedriver file path here
chrome_path = 'C:/Program Files/Google/Chrome/Application/chromedriver.exe'
# !change the output file path here
outputFile = 'D:/git/data_crawl/raw_data/final_Trademark.xlsx'



def read_all_ids(id_path):
    ids = []
    xl = pd.read_excel(id_path, index_col=None, header=None)
    ids = xl[0].values.tolist()
    return ids

def get_proxyList():
    url = 'http://dps.kdlapi.com/api/getdps/?orderid=911955699122279&num=2&pt=1&format=json&sep=1'
    resp = requests.get(url)
    json_data = resp.json()
    proxylist = json_data['data']['proxy_list']
    return proxylist


def generate_resultsUrlList(ids):
    resultsUrlList = [
        'https://trademarks.ipo.gov.uk/ipo-tmcase/page/Results/1/'+id for id in ids]
    return resultsUrlList


def get_tabs(url, proxylist):
    for PROXY in proxylist:
        try:
            chrome_options = Options()
            chrome_options.add_argument('--headless')
            chrome_options.add_argument('--proxy-server={}'.format(PROXY))
            driver = webdriver.Chrome(
                executable_path=chrome_path, options=chrome_options)
            tabContents = []
            driver.get(url)
            webPage = driver.page_source
            noSuchId = driver.find_elements_by_xpath('/html/body/main/form/div[3]/button')
            if len(noSuchId) > 0:
                return 'Search for a trade mark'
            print('1' + driver.current_url)
            soup = BeautifulSoup(webPage, 'lxml')
            tabs = soup.select(
                'body>main>.tab-container.ui-tabs.ui-corner-all.ui-widget.ui-widget-content>div')
            print(len(tabs))
            if len(tabs) == 0:
                driver.delete_all_cookies()
                driver.quit()
                continue
            resultContent = get_results(tabs)
            historyEle = driver.find_element_by_xpath(
                '/html/body/main/div[4]/div/p[1]/a')
            if historyEle is not None:
                historyUrl = historyEle.get_attribute('href')
                driver.get(historyUrl)
                print('2' + driver.current_url)
                webPage = driver.page_source
                soup = BeautifulSoup(webPage, 'lxml')
                tabs = soup.select(
                    'body>main>.tab-container.ui-tabs.ui-corner-all.ui-widget.ui-widget-content>div')
                print(len(tabs))
                if len(tabs) == 0:
                    driver.delete_all_cookies()
                    driver.quit()
                    continue
                historyContent = get_history(tabs)
            driver.delete_all_cookies()
            driver.quit()
            tabContents = resultContent + historyContent
            return tabContents
        except Exception as e:
            print(e)
            driver.delete_all_cookies()
            driver.quit()
            continue


def get_results(tabs):
    resultContent = []
    overview = ['', '', '']
    goods = ['']
    names = ['']
    publications = ['', '']
    for i in range(1, len(tabs)):
        tab = tabs[i]
        title = tab.find('h2').get_text().strip()
        if title == 'Overview':
            overview = get_overview(tab)
        elif title == 'Goods and services':
            goods = get_goods(tab)
        elif title == 'Names and addresses':
            names = get_names(tab)
        elif title == 'Publications':
            publications = get_publications(tab)
    resultContent = overview + goods + names + publications
    return resultContent


def get_overview(tab):
    filingDate = ''
    registerDate = ''
    renewalDate = ''
    dates = tab.select('.grid-row>dl')
    for date in dates:
        key = date.find('dt').get_text().strip()
        value = date.find('dd').get_text().strip()
        if key.startswith('Filing'):
            filingDate = value
        elif key.startswith('Date'):
            registerDate = value
        elif key.startswith('Renewal'):
            renewalDate = value
    return [filingDate, registerDate, renewalDate]


def get_goods(tab):
    classes = tab.select(
        '.accordion.with-js>.subsection-wrapper>.accordion-section')
    classNames = ''
    for cl in classes:
        classNames += cl.find('h2').get_text().strip() + \
            '\n' + '--------------' + '\n'
    return [classNames]


def get_names(tab):
    names = ''
    items = tab.find_all(['h3', 'dt', 'dd'])
    for item in items:
        names += item.get_text().strip() + '\n'
    return[names]


def get_publications(tab):
    dls = tab.find_all('dl')
    journal = ''
    publicationDate = ''
    for dl in dls:
        if dl.find('dt').get_text() == 'Journal':
            journal = dl.find('dd').get_text().strip()
        elif dl.find('dt').get_text() == 'Date of publication':
            publicationDate = dl.find('dd').get_text().strip()
    return [journal, publicationDate]


def get_history(tabs):
    historyContent = []
    status = ['']
    events = ['']
    goods_history = ['', '']
    archived = ['']
    for i in range(1, len(tabs)):
        tab = tabs[i]
        title = tab.find('h2').get_text().strip()
        if title == 'Status history':
            status = get_status(tab)
        elif title == 'Event history':
            events = get_events(tab)
        elif title == 'Goods and services history':
            goods_history = get_goods_history(tab)
        elif title == 'Archived history':
            archived = get_archived(tab)
    historyContent = status + events + goods_history + archived
    return historyContent


def get_status(tab):
    status = ''
    dates = tab.select('.grid-row')
    for date in dates:
        ds = date.find_all('dl')
        for d in ds:
            status += d.get_text().strip() + ',          '
        status += '\n' + '--------------' + '\n'
    return [status]


def get_events(tab):
    events = ''
    eves = tab.select('.grid-row')
    for eve in eves:
        dls = eve.find_all('dl')
        for dl in dls:
            key = dl.find('dt').get_text().strip()
            val = dl.find('dd').get_text().strip()
            events += key + ':        ' + val + '\n'
        events += '\n' + '--------------' + '\n'
    return [events]


def get_goods_history(tab):
    amendedDate = ''
    amended = tab.find('dl')
    amendedDate = amended.find('dd').get_text()
    classes = tab.select(
        '.accordion.with-js>.subsection-wrapper>.accordion-section')
    classNames = ''
    for cl in classes:
        classNames += cl.find('h2').get_text().strip() + \
            '\n' + '--------------' + '\n'
    return [amendedDate, classNames]


def get_archived(tab):
    link = tab.find('a').get('href')
    return [link]


def write_results():
    xlsFile = openpyxl.Workbook()
    sheet1 = xlsFile.create_sheet(index=0)
    header = ['ID', 'Filing date', 'Date of entry in register', 'Renewal date',
              'Classes and terms', 'Names and addresses',
              'First advert - Journal', 'Date of publication', 'Status history', 'Event history',
              'Goods and services history - Date amended', 'Goods and services history - Classes and terms',
              'Archived history - PDF link']
    for i in range(1, len(header)+1):
        sheet1.cell(1, i).value = header[i-1]
    results = dict()
    ids = read_all_ids(id_path)
    resultsUrlList = generate_resultsUrlList(ids)
    proxylist = get_proxyList()
    start = time.time()
    n = 0
    for i in range(0, len(resultsUrlList)):
        resultsUrl = resultsUrlList[i]
        urlId = ids[i]
        middle = time.time()
        if middle - start > 120:
            proxylist = get_proxyList()
            start = time.time()
        resultsTabContents = get_tabs(resultsUrl, proxylist)
        if type(resultsTabContents) == str:
            print('Invalid ID: {}'.format(urlId))
            n += 1
            continue
        elif resultsTabContents == None:
            proxylist = get_proxyList()
            resultsTabContents = get_tabs(resultsUrl, proxylist)
        results[urlId] = [urlId] + resultsTabContents
        for j in range(1, len(header)+1):
            sheet1.cell(i+2-n, j).value = results[urlId][j-1]
        xlsFile.save(outputFile)
    return results


def run():
    results = write_results()


if __name__ == '__main__':
    run()