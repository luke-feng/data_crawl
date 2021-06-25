# coding: utf-8

import time
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.support import expected_conditions as Expect
from selenium.webdriver.support import wait
from selenium.webdriver.support.ui import Select
from selenium.webdriver.support.wait import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support.ui import WebDriverWait as Wait
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.by import By
from selenium import webdriver
import requests
from bs4 import BeautifulSoup
import re
import tqdm
import random
import sys
import os
import xlwt
import io
import sys
import urllib.request
import openpyxl
from concurrent.futures import ThreadPoolExecutor
import concurrent.futures
import datetime

# chrome_options = Options()
# chrome_options.add_argument('--headless')
# chrome_options.add_experimental_option('excludeSwitches', ['enable-automation'])
# # Change your chrome path here
# chrome_path = 'C:/Program Files/Google/Chrome/Application/chromedriver.exe'
# driver = webdriver.Chrome(
#     executable_path=chrome_path, options=chrome_options)
# qstart = '01/06/2021'
# qend = '03/06/2021'
# driver.get('https://planningsearch.harrow.gov.uk/planning/search-applications')
# a = driver.page_source
# time.sleep(1)
# fromdate = driver.find_element_by_xpath('/html/body/div[1]/div[2]/div/div/div[1]/div[1]/div/div[6]/div/div[2]/input')

# fromdate.send_keys(qstart)
# enddate = driver.find_element_by_xpath('/html/body/div[1]/div[2]/div/div/div[1]/div[1]/div/div[7]/div/div[2]/input')
# enddate.send_keys(qend)
# search = driver.find_element_by_xpath('/html/body/div[1]/div[2]/div/div/div[1]/div[1]/div/div[14]/p/button')
# print(search.text)
# search.click()
# time.sleep(1)
# trs = driver.find_elements_by_class_name('civica-keyobject-basicdetails')
# print(len(trs))
# tds = trs[0]
# name = tds.find_element_by_css_selector('a').text
# link = tds.find_element_by_css_selector('a').get_attribute('href')
# refNo = name.split('-')[1]
# valid = name.split('-')[2]
# divs = trs[0].find_elements_by_css_selector('div')
# address= divs[0].text
# prop = divs[2].text
# print(name, refNo,link,valid, address )
# link = 'https://planningsearch.harrow.gov.uk/planning/search-applications#VIEW?RefType=GFPlanning&KeyNo=967251&KeyText=Subject'
# driver.get(link)
# time.sleep(5)
# frame = driver.find_element_by_xpath('/html/body/div/div[2]/div/div/div[2]/div[1]/div/div/div[2]/div/div[3]/div/div')
# print(frame)
# lis = frame.find_elements_by_class_name('civicadetail')
# print(len(lis))
# for li in lis:
#     kv = li.find_elements_by_css_selector('div')
#     key = kv[0].get_attribute('textContent')
#     value = kv[1].get_attribute('textContent')
#     print(key, value)

# rpb5s = driver.find_elements_by_class_name('row.pad-bottom-5')
# for rpb5 in  rpb5s:
#     md5 = rpb5.find_element_by_class_name('col-md-5').text
#     md7 = rpb5.find_element_by_class_name('col-md-7').text
#     print(md5, md7)
# t3 = time.time()
# print(t3-t2)

class Task5:
    def __init__(self, resultPath, chrome_path):
        self.finalResult = []
        self.resultPath = resultPath
        chrome_options = Options()
        self.line = 0
        self.xlsFile = openpyxl.Workbook()
        self.sheet1 = self.xlsFile.create_sheet(index=0)
        chrome_options.add_argument('--headless')
        self.max_workers=10
        self.executor = ThreadPoolExecutor(max_workers=self.max_workers)
        # for main thread
        self.driver = webdriver.Chrome(
            executable_path=chrome_path, options=chrome_options)
        # for workers
        self.driver1 = webdriver.Chrome(
            executable_path=chrome_path, options=chrome_options)
        self.driver2 = webdriver.Chrome(
            executable_path=chrome_path, options=chrome_options)
        self.driver3 = webdriver.Chrome(
            executable_path=chrome_path, options=chrome_options)
        self.driver4 = webdriver.Chrome(
            executable_path=chrome_path, options=chrome_options)
        self.driver5 = webdriver.Chrome(
            executable_path=chrome_path, options=chrome_options)
        self.driver6 = webdriver.Chrome(
            executable_path=chrome_path, options=chrome_options)
        self.driver7 = webdriver.Chrome(
            executable_path=chrome_path, options=chrome_options)
        self.driver8 = webdriver.Chrome(
            executable_path=chrome_path, options=chrome_options)
        self.driver9 = webdriver.Chrome(
            executable_path=chrome_path, options=chrome_options)
        self.driver10 = webdriver.Chrome(
            executable_path=chrome_path, options=chrome_options)
        self.workers = [self.driver1, self.driver2,
                        self.driver3, self.driver4, self.driver5,self.driver6, self.driver7,
                        self.driver8, self.driver9, self.driver10]

    def get_search_results_page(self, url, qstart, qend, council):
        self.driver.get(url)
        Wait(self.driver, 10).until(
            Expect.presence_of_element_located(
                (By.XPATH, "/html/body/div[1]/div[2]/div/div/div[1]/div[1]/div/div[6]/div/div[2]/input"))
        )
        fromdate = self.driver.find_element_by_xpath('/html/body/div[1]/div[2]/div/div/div[1]/div[1]/div/div[6]/div/div[2]/input')
        fromdate.send_keys(qstart)
        enddate = self.driver.find_element_by_xpath('/html/body/div[1]/div[2]/div/div/div[1]/div[1]/div/div[7]/div/div[2]/input')
        enddate.send_keys(qend)
        search = self.driver.find_element_by_xpath('/html/body/div[1]/div[2]/div/div/div[1]/div[1]/div/div[14]/p/button')
        search.click()
        print(qstart, qend, council)
        Wait(self.driver, 10).until(
            Expect.presence_of_element_located(
                (By.XPATH, "/html/body/div[1]/div[2]/div/div/div[1]/div[2]/div/div/div[1]/div[2]/ul/li[1]"))
        )
        self.driver.find_elements_by_class_name('civica-keyobject-basicdetails')

    def run(self, url, council):
        header = ['Council', 'Ref. No', 'Name', 'Address', 'Received',
                    'Validated', 'Status', 'MetaInfo', 'Link',
                    'Applicant_Name', 'Premises_Address', 'Postcode', 'Ward',
                    'Date_Received', 'Case_Officer', 'Application_Type', 'Agent_Name',
                    'Agent_Address', 'Proposal', 'Registered_Date', 'Application', 
                    'Appeal_Date', 'Decision_Date', 'Decision', 'Appeal_Decision_Date'
                  ]
        startday = ['01/01', '01/02', '01/03', '01/04', '01/05', '01/06',
                    '01/07', '01/08', '01/09', '01/10', '01/11', '01/12']
        endday = ['01/02', '01/03', '01/04', '01/05', '01/06', '01/07',
                  '01/08', '01/09', '01/10', '01/11', '01/12', '31/12']
        for col in range(1, len(header)+1):
            self.sheet1.cell(1, col).value = header[col-1]
        
        qstart = datetime.date(2000,1,1)
        last = datetime.date(2020,12,31)
        qend = qstart
        while qend < last:
            qend = qstart + datetime.timedelta(days=30)
            if qend >= last:
                qend = last
            # par = tqdm.tqdm(ncols=100)
            sstart = qstart.strftime("%d/%m/%Y")
            send = qend.strftime("%d/%m/%Y")
            qstart = qstart + datetime.timedelta(days=31)  
            self.get_search_results_page(url, sstart, send, council)
            par = tqdm.tqdm(ncols=100)

            try:
                while True:                        
                    searchResults = self.driver.find_elements_by_class_name('civica-keyobject-basicdetails')
                    # print('searchResults', len(searchResults))
                    next = self.driver.find_elements_by_xpath('/html/body/div[1]/div[2]/div/div/div[1]/div[2]/div/div/div[2]/div/div/div')                
                    # print('next', len(next))
                    if len(next) > 0:
                        start = []
                        end = []
                        for i in range(0, len(searchResults), self.max_workers):
                            start.append(i)
                            if i + self.max_workers >= len(searchResults):
                                end.append(len(searchResults))
                            else:
                                end.append(i+self.max_workers)
                        for i, s in enumerate(start):
                            self.asyn_page(
                                url_list=searchResults[start[i]: end[i]], council=council)
                            par.update(self.max_workers)
                        if len(next) == 1:
                            next_button = next[0]
                            next_button.click()
                            Wait(self.driver, 10).until(
                                Expect.presence_of_element_located(
                                    (By.XPATH, "/html/body/div[1]/div[2]/div/div/div[1]/div[2]/div/div/div[1]/div[2]/ul/li[1]"))
                            )
                        if len(next) == 2:
                            next_button = next[1]
                            next_button.click()
                            Wait(self.driver, 10).until(
                                Expect.presence_of_element_located(
                                    (By.XPATH, "/html/body/div[1]/div[2]/div/div/div[1]/div[2]/div/div/div[1]/div[2]/ul/li[1]"))
                            )
                        if len(next) == 3:
                            next_button = next[2]
                            if 'disabled-btn' not in  next_button.get_attribute('class'):
                                next_button.click()
                                Wait(self.driver, 10).until(
                                    Expect.presence_of_element_located(
                                        (By.XPATH, "/html/body/div[1]/div[2]/div/div/div[1]/div[2]/div/div/div[1]/div[2]/ul/li[1]"))
                                )
                            else:
                                break
                    else:
                        start = []
                        end = []
                        for i in range(0, len(searchResults), self.max_workers):
                            start.append(i)
                            if i + self.max_workers >= len(searchResults):
                                end.append(len(searchResults))
                            else:
                                end.append(i+self.max_workers)
                        for i, s in enumerate(start):
                            self.asyn_page(
                                url_list=searchResults[start[i]:end[i]], council=council)
                            par.update(self.max_workers) 
                        print('break')                   
                        break
            except Exception as exc:
                print(qstart, qend, exc)
                continue
            par.close()
        par = tqdm.tqdm(total = len(self.finalResult), ncols=100)
        for data in self.finalResult:
            par.update(1)
            for col in range(1, len(data)+1):
                self.sheet1.cell(self.line+2, col).value = data[col-1]            
            self.line += 1
        self.xlsFile.save(self.resultPath)
        par.close()
        self.xlsFile.close()
        print('total page is {}'.format(len(self.finalResult)))
        print('get webpage finish!')

    def __del__(self):
        self.driver.close()
        self.driver1.close()
        self.driver2.close()
        self.driver3.close()
        self.driver4.close()
        self.driver5.close()
        self.driver6.close()
        self.driver7.close()
        self.driver8.close()
        self.driver9.close()
        self.driver10.close()
        print('>>>>[Well Done]')

    def get_information(self, browser, result, council):
        tds = result
        name = tds.find_element_by_css_selector('a').text
        link = tds.find_element_by_css_selector('a').get_attribute('href')
        refNo = name.split('-')[1]
        Validated = name.split('-')[2]
        divs = result.find_elements_by_css_selector('div')
        address= divs[0].text
        Received= ''
        Status= ''
        metaInfo = divs[2].get_attribute('textContent')
        details = []
        details = self.get_details(browser, link)
        res = [council,refNo, name, address, Received,
            Validated, Status, metaInfo, link] + details
        return res

    def get_details(self, browser, url):
        details = []
        Applicant_Name = ''  
        Premises_Address = ''
        Postcode = ''
        Ward = ''
        Date_Received = ''
        Case_Officer=''
        Application_Type = ''
        Agent_Name = ''
        Agent_Address=''
        Proposal = ''
        Registered_Date = ''
        Application = ''
        Appeal_Date = ''
        Decision_Date = ''
        Decision = ''
        Appeal_Decision_Date = ''
        browser.get(url)
        Wait(browser, 20).until(
            Expect.presence_of_element_located(
                (By.XPATH, '/html/body/div[1]/div[2]/div/div/div[2]/div/div/div/div[2]/div/div[3]/div/div/div[1]/div[1]'))
        )  
        frame = browser.find_element_by_xpath('/html/body/div/div[2]/div/div/div[2]/div[1]/div/div/div[2]/div/div[3]/div/div')
        rpb5s = frame.find_elements_by_class_name('civicadetail')
        if len(rpb5s)>0:
            for rpb5 in  rpb5s:
                kv = rpb5.find_elements_by_css_selector('div')
                key = kv[0].get_attribute('textContent').strip()
                
                value = kv[1].get_attribute('textContent').strip()
                # print(value)
                
                if key == 'Applicant Name':
                    Applicant_Name = value
                elif key == 'Premises Address':
                    Premises_Address = value
                elif key == 'Postcode':
                    Postcode = value
                elif key == 'Ward':
                    Ward = value
                elif key == 'Date Received':
                    Date_Received = value
                elif key == 'Case Officer':
                    Case_Officer = value
                elif key == 'Application Type':
                    Application_Type = value
                elif key == 'Agent Name':
                    Agent_Name = value
                elif key == 'Agent Address':
                    Agent_Address = value
                elif key == 'Proposal':
                    Proposal = value
                elif key == 'Registered Date':
                    Registered_Date = value
                elif key == 'Application':
                    Application = value
                elif key == 'Appeal Date':
                    Appeal_Date = value
                elif key == 'Decision Date':
                    Decision_Date = value
                elif key == 'Decision':
                    Decision = value
                elif key == 'Appeal Decision Date':
                    Appeal_Decision_Date = value
               
        details = [Applicant_Name, Premises_Address, Postcode, Ward,
        Date_Received, Case_Officer, Application_Type, Agent_Name,
        Agent_Address, Proposal, Registered_Date, Application, 
        Appeal_Date, Decision_Date, Decision, Appeal_Decision_Date]
        return details

    def asyn_page(self, url_list, council):
        future_to_url  = dict()
        for i, url in enumerate(url_list):
            t = self.executor.submit(self.get_information,
                                 browser=self.workers[i], result=url_list[i], council = council)
            future_to_url[t] = url               
        for future in concurrent.futures.as_completed(future_to_url):
            url = future_to_url[future]
            try:
                data = future.result()
                if data != None:
                    self.finalResult.append(data)
            except Exception as exc:
                print('%r generated an exception: %s' % (url, exc))

if __name__ == '__main__':
    # chage the data path here
    datapath = 'D:/git/data_crawl/raw_data/gazette/'
    chrome_path = 'C:/Program Files/Google/Chrome/Application/chromedriver.exe'
    urllist = [
        'https://planningsearch.harrow.gov.uk/planning/search-applications'
    ]
    councillist = [
        'London Borough of Harrow'

    ]
    for i in range(0, len(urllist)):
        infoPageName = datapath + councillist[i] + '.xlsx'
        url = urllist[i]
        task = Task5(infoPageName, chrome_path)
        task.run(url,councillist[i])
        # try: 
        #     infoPageName = datapath + councillist[i] + '.xlsx'
        #     url = urllist[i]
        #     task = Task5(infoPageName, chrome_path)
        #     task.run(url,councillist[i])
        # except Exception as exc:
        #     print(exc)
