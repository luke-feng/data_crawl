import time
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.support import expected_conditions as Expect
from selenium.webdriver.support import wait
from selenium.webdriver.support.ui import Select
from selenium.webdriver.support.wait import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support.ui import WebDriverWait as Wait
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.by import By
from selenium import webdriver
import requests
from bs4 import BeautifulSoup
import re
import tqdm
import random
import sys
import os
import xlwt
import io
import sys
import urllib.request
import openpyxl
from concurrent.futures import ThreadPoolExecutor
import concurrent.futures
import datetime
import pandas as pd

def get_all_links():
    chrome_options = Options()
    chrome_options.add_argument('--headless')
    # Change your chrome path here
    chrome_path = 'C:/Program Files/Google/Chrome/Application/chromedriver.exe'
    driver = webdriver.Chrome(
        executable_path=chrome_path, options=chrome_options)


    select_values = ['CA', 'CH', 'HC', 'IB', 'OB', 'WA']
    driver.get('https://planning-lbhounslow.msappproxy.net/Planning_Search_Advanced.aspx')
    cookie = driver.find_element_by_xpath('/html/body/main/div/div[1]/article/form/div[3]/p[3]/input').click()
    planning = driver.find_element_by_xpath('/html/body/main/div/div[1]/article/div/div/div/a/h2').click()
    adv_planning = driver.find_element_by_xpath('/html/body/main/div/div[1]/article/form/div[3]/ul/li[2]/a').click()
    datapath = 'D:/git/data_crawl/raw_data/gazette/'

    with open(datapath+'p7.csv', 'w') as linkfile:
        for sv in select_values:
            driver.get('https://planning-lbhounslow.msappproxy.net/Planning_Search_Advanced.aspx')
            search_area = driver.find_element_by_xpath('/html/body/main/div/div[1]/article/form/div[4]/div[4]/p/select')
            s1 = Select(search_area)
            s1.select_by_value(sv)
            qstart = '01/01/2000'
            qend = '01/02/2021'
            fromdate = driver.find_element_by_xpath('/html/body/main/div/div[1]/article/form/div[4]/div[5]/p[7]/input')
            fromdate.send_keys(qstart)
            enddate = driver.find_element_by_xpath('/html/body/main/div/div[1]/article/form/div[4]/div[5]/p[8]/input')
            enddate.send_keys(qend)
            search = driver.find_element_by_xpath('/html/body/main/div/div[1]/article/form/div[4]/p[2]/input[2]')
            search.click()
            tat = int(driver.find_element_by_xpath('/html/body/main/div/div[1]/article/form/p[2]/span').text)
            par = tqdm.tqdm(total=tat, ncols=100)
            while True:
                next = driver.find_elements_by_xpath('/html/body/main/div/div[1]/article/form/div[4]/div/div[3]/a[1]')
                if len(next) > 0:                
                    results = driver.find_elements_by_xpath('/html/body/main/div/div[1]/article/form/p/strong/a')
                    for res in results:
                        par.update(1)
                        ref = res.text
                        link = res.get_attribute('href')
                        linkfile.write(ref+','+link+'\n')
                    next[0].click()
                else:
                    results = driver.find_elements_by_xpath('/html/body/main/div/div[1]/article/form/p/strong/a')
                    for res in results:
                        par.update(1)
                        ref = res.text
                        link = res.get_attribute('href')
                        linkfile.write(ref+','+link+'\n')
                    break
    driver.close()
    linkfile.close()

class Task5:
    def __init__(self, resultPath, chrome_path, linkPath):
        self.finalResult = []
        self.linkpath = linkPath
        self.resultPath = resultPath
        chrome_options = Options()
        chrome_options.add_argument('--headless')
        chrome_options.add_experimental_option ('excludeSwitches', ['enable-logging'])
        self.line = 0
        self.xlsFile = openpyxl.Workbook()
        self.sheet1 = self.xlsFile.create_sheet(index=0)
        self.max_workers = 16
        self.k = []
        self.executor = ThreadPoolExecutor(max_workers=self.max_workers)
        # for main thread
        self.driver = webdriver.Chrome(
            executable_path=chrome_path, options=chrome_options)
        # for workers
        self.driver1 = webdriver.Chrome(
            executable_path=chrome_path, options=chrome_options)
        self.driver2 = webdriver.Chrome(
            executable_path=chrome_path, options=chrome_options)
        self.driver3 = webdriver.Chrome(
            executable_path=chrome_path, options=chrome_options)
        self.driver4 = webdriver.Chrome(
            executable_path=chrome_path, options=chrome_options)
        self.driver5 = webdriver.Chrome(
            executable_path=chrome_path, options=chrome_options)
        self.driver6 = webdriver.Chrome(
            executable_path=chrome_path, options=chrome_options)
        self.driver7 = webdriver.Chrome(
            executable_path=chrome_path, options=chrome_options)
        self.driver8 = webdriver.Chrome(
            executable_path=chrome_path, options=chrome_options)
        self.driver9 = webdriver.Chrome(
            executable_path=chrome_path, options=chrome_options)
        self.driver10 = webdriver.Chrome(
            executable_path=chrome_path, options=chrome_options)
        self.driver11 = webdriver.Chrome(
            executable_path=chrome_path, options=chrome_options)
        self.driver12 = webdriver.Chrome(
            executable_path=chrome_path, options=chrome_options)
        self.driver13 = webdriver.Chrome(
            executable_path=chrome_path, options=chrome_options)
        self.driver14 = webdriver.Chrome(
            executable_path=chrome_path, options=chrome_options)
        self.driver15 = webdriver.Chrome(
            executable_path=chrome_path, options=chrome_options)
        self.workers = [self.driver,
                        self.driver1, self.driver2,
                        self.driver3, self.driver4, 
                        self.driver5,
                        self.driver6, self.driver7, self.driver8, 
                        self.driver9, self.driver10,
                        self.driver11, self.driver12,
                        self.driver13, self.driver14, 
                        self.driver15,
                        ]
    
    def get_proxyList(self):
        url = 'http://dps.kdlapi.com/api/getdps/?orderid=901955512630336&num=1&pt=1&sep=1'
        resp = requests.get(url)
        json_data = resp.json()
        proxylist = json_data['data']['proxy_list']
        return proxylist

    def read_all_links(self, linkpath):
        links =pd.read_csv(linkpath, index_col=None, header=None, engine='python' , error_bad_lines=False).values
        return links
    
    def click_cookie(self, url, driver):
        driver.get(url)
        cookie = driver.find_elements_by_xpath('/html/body/main/div/div[1]/article/form/div[3]/p[3]/input')
        # print('cookie', len(cookie))
        print(cookie[0].get_attribute('value'))
        cookie[0].click()


    def run(self, url, council):

        for driver in self.workers:
            self.click_cookie( url, driver)

        header = ['council', 'refNo', 'link',
                    'Address', 'Proposal','Application_Type',  'Status', 'Date',
                    'System_Reference', 'Planning_Reference', 'Ward', 'Planning_officer', 'Application_Received',
                    'Application_Accepted', 'Initial_Status', 'Decision', 'Decision_Issued', 'Expiry_Date'
                  ]

        for col in range(1, len(header)+1):
            self.sheet1.cell(1, col).value = header[col-1]

        urls = self.read_all_links(self.linkpath)
        start = []
        end = []
        par = tqdm.tqdm(total = len(urls), ncols=100)
        
        for i in range(0, len(urls), self.max_workers):
            start.append(i)
            if i + self.max_workers >= len(urls):
                end.append(len(urls))
            else:
                end.append(i+self.max_workers)
        t1 = time.time()
        for i, s in enumerate(start):
            try: 
                
                self.asyn_page(
                    url_list=urls[start[i]: end[i]], council=council)
                t2 = time.time()
                par.update(self.max_workers)
                steps = t2 - t1
                if steps >= 3600:
                    t1 = time.time()
                    for driver in self.workers:
                        self.click_cookie( url, driver)
                
            except Exception as exc:
                print(exc)
                continue   
        par.close()

        par = tqdm.tqdm(total = len(self.finalResult), ncols=100)
        for data in self.finalResult:
            par.update(1)
            for col in range(1, len(data)+1):
                self.sheet1.cell(self.line+2, col).value = data[col-1]            
            self.line += 1
        self.xlsFile.save(self.resultPath)
        par.close()

    def __del__(self):
        self.driver.close()
        self.driver1.close()
        self.driver2.close()
        self.driver3.close()
        self.driver4.close()
        self.driver5.close()
        self.driver6.close()
        self.driver7.close()
        self.driver8.close()
        self.driver9.close()
        self.driver10.close()
        self.driver11.close()
        self.driver12.close()
        self.driver13.close()
        self.driver14.close()
        self.driver15.close()
        print('>>>>[Well Done]')

    def get_information(self, browser, result, council):
        link = result[1]
        refNo = result[0]        
        details = []        
        details = self.get_details(browser, link)
        
        res = [council, refNo, link] + details 
        return res

    

    def get_details(self, browser, url):
        details = []
        time.sleep(1)
        browser.get(url)
        cookie = browser.find_elements_by_xpath('/html/body/main/div/div[1]/article/form/div[3]/p[3]/input')
        if len(cookie) > 0:
            print(cookie[0].get_attribute('value'))
            cookie[0].click()
            browser.get(url)
        i = 0
        while i <= 10:
            i += 1
            time.sleep(1)
            Addresses =browser.find_elements_by_id('MainContent_lbl_site_description')
            if len(Addresses) >  0:
                break
            else:
                browser.get(url)
        Application_Type=''
        Status=''
        Date=''
        System_Reference=''
        Planning_Reference=''
        Ward=''
        Planning_officer=''
        Application_Received=''
        Application_Accepted=''
        Initial_Status=''
        Decision=''
        Decision_Issued=''
        Expiry_Date=''
        Address = ''
        Proposal = ''
        try:
            Address =browser.find_element_by_id('MainContent_lbl_site_description').text
            Proposal =browser.find_element_by_id('MainContent_lbl_Proposal').text
            keys = browser.find_elements_by_xpath('/html/body/main/div/div[1]/article/form/dl/dt')
            values = browser.find_elements_by_xpath('/html/body/main/div/div[1]/article/form/dl/dd')
            for i, key in enumerate(keys):
                key = key.text.strip()
                if key not in self.k:
                    self.k.append(key)
                content = values[i].text.strip()
                if key == 'Application Type':
                    Application_Type = content
                elif key == 'Status':
                    Status = content
                elif key == 'Date':
                    Date = content
                elif key == 'System Reference':
                    System_Reference = content
                elif key == 'Planning Reference':
                    Planning_Reference = content
                elif key == 'Ward':
                    Ward = content
                elif key == 'Planning officer':
                    Planning_officer = content
                elif key == 'Application Received':
                    Application_Received = content
                elif key == 'Application Accepted':
                    Application_Accepted = content
                elif key == 'Initial Status':
                    Initial_Status = content
                elif key == 'Decision':
                    Decision = content
                elif key == 'Decision Issued':
                    Decision_Issued = content
                elif key == 'Expiry Date':
                    Expiry_Date = content
        except Exception as exc:
            print(url, exc)

        details = [Address, Proposal,Application_Type,  Status, Date,
        System_Reference, Planning_Reference, Ward, Planning_officer, Application_Received,
        Application_Accepted, Initial_Status, Decision, Decision_Issued, Expiry_Date
        ]
        return details

    

    def asyn_page(self, url_list, council):
        future_to_url  = dict()
        for i, url in enumerate(url_list):
            t = self.executor.submit(self.get_information,
                                 browser=self.workers[i], result=url_list[i], council = council)
            future_to_url[t] = url               
        for future in concurrent.futures.as_completed(future_to_url):
            url = future_to_url[future]
            try:
                data = future.result()
                if data != None:
                    self.finalResult.append(data)
            except Exception as exc:
                print('%r generated an exception: %s' % (url, exc))


if __name__ == '__main__':
    # chage the data path here
    datapath = 'D:/git/data_crawl/raw_data/gazette/'
    chrome_path = 'C:/Program Files/Google/Chrome/Application/chromedriver.exe'
    linkPath = datapath + 'p7.csv'
    urllist = [
        'https://planning-lbhounslow.msappproxy.net/Planning_Search_Advanced.aspx'
    ]
    councillist = [
        'London Borough of Hounslow'
    ]

    for i in range(0, len(urllist)):
        try: 
            infoPageName = datapath + councillist[i] + '.xlsx'
            url = urllist[i]
            task = Task5(infoPageName, chrome_path, linkPath)
            task.run(url,councillist[i])
            print(task.k)
        except Exception as exc:
            print(exc)
    
