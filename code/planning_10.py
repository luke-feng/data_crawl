# coding: utf-8

import time
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.support import expected_conditions as Expect
from selenium.webdriver.support import wait
from selenium.webdriver.support.ui import Select
from selenium.webdriver.support.wait import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support.ui import WebDriverWait as Wait
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.by import By
from selenium import webdriver
import requests
from bs4 import BeautifulSoup
import re
import tqdm
import random
import sys
import os
import xlwt
import io
import sys
import urllib.request
import openpyxl
from concurrent.futures import ThreadPoolExecutor
import concurrent.futures
import datetime
import urllib3
urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)

class Task5:
    def __init__(self, resultPath):
        self.finalResult = []
        self.resultPath = resultPath
        self.line = 0
        self.xlsFile = openpyxl.Workbook()
        self.sheet1 = self.xlsFile.create_sheet(index=0)
        self.header = {
            'Accept': 'application/json, text/plain, */*',
            'Accept-Encoding': 'gzip, deflate, br',
            'Accept-Language': 'en-GB,en;q=0.9,en-US;q=0.8,zh-CN;q=0.7,zh;q=0.6',
            'Connection': 'keep-alive',
            'Host': 'planningapi.agileapplications.co.uk',
            'Origin': 'https://planning.redbridge.gov.uk',
            'Referer': 'https://planning.redbridge.gov.uk/',
            'sec-ch-ua': '" Not;A Brand";v="99", "Microsoft Edge";v="91", "Chromium";v="91"',
            'sec-ch-ua-mobile': '?0',
            'Sec-Fetch-Dest': 'empty',
            'Sec-Fetch-Mode': 'cors',
            'Sec-Fetch-Site': 'cross-site',
            'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/91.0.4472.114 Safari/537.36 Edg/91.0.864.54',
            'x-client': 'RG',
            'x-product': 'CITIZENPORTAL',
            'x-service': 'PA'
        }

    def run(self, council):
        header = ['Council', ' Link', 'id', 'reference', 'proposal', 'location', 'username', 
                  'applicantSurname', 'registrationDate', 'decisionDate', 'decisionText', 'finalGrantDate',
                   'extensionDate', 'appealLodgedDate', 'appealDecisionDate', 'abpReference', 'appealDecision', 
                   'postcode', 'easting', 'northing', 'area', 'ward', 
                   'parish', 'locationURL', 'fullProposal', 'registerDate', 'webReference', 
                   'dispatchDate', 'statusOwner', 'statusDescriptionOwner', 'statusNonOwner', 'statusDescriptionNonOwner', 
                   'applicationTypeId', 'applicationType', 'statutoryExpiryDate', 'decisionExpiryDate', 'agentSurname', 
                   'officerName', 'officerTelephone', 'officerEmail', 'appealType', 'receivedDate', 
                   'article35', 'commentsMode', 'publicityEndDate', 'tracked', 'developmentCategory', 
                   'applicationDate', 'decisionDueDate', 'pressNoticeStartDate', 'validDate', 'uprn',
                    'agentName', 'agentTitle', 'agentInitials']
        for col in range(1, len(header)+1):
            self.sheet1.cell(1, col).value = header[col-1]
        
        qstart = datetime.date(2000,1,1)
        last = datetime.date(2020,12,31)
        qend = qstart
        while qend < last:
            qend = qstart + datetime.timedelta(days=100)
            if qend >= last:
                qend = last
            # par = tqdm.tqdm(ncols=100)
            sstart = qstart.strftime("%Y-%m-%d")
            send = qend.strftime("%Y-%m-%d")
            qstart = qstart + datetime.timedelta(days=101) 
            url_q = 'https://planningapi.agileapplications.co.uk/api/application/search?decisionDateFrom={}&decisionDateTo={}'.format(sstart, send)
            page = requests.get(url_q,verify=False, headers=self.header).json()
            print(council, sstart, send)
            total = int(page['total'])
            par = tqdm.tqdm(total=total, ncols=100)
            results = page['results']
            for r in results:
                id = r['id']          
                url_d = 'https://planningapi.agileapplications.co.uk/api/application/{}'.format(str(id))
                page = requests.get(url_d,verify=False, headers=self.header).json()
                res = list(page.values())
                result = [council, url_d] + res
                self.finalResult.append(result)
                par.update(1)
            par.close()

 
        par = tqdm.tqdm(total = len(self.finalResult), ncols=100)
        for data in self.finalResult:
            par.update(1)
            try:
                for col in range(1, len(data)+1):
                    self.sheet1.cell(self.line+2, col).value = data[col-1]            
                self.line += 1
            except:
                continue
        self.xlsFile.save(self.resultPath)
        par.close()
        self.xlsFile.close()
        print('total page is {}'.format(len(self.finalResult)))
        print('get webpage finish!')

    def __del__(self):
        print('>>>>[Well Done]')

    
if __name__ == '__main__':
    # chage the data path here
    datapath = 'D:/git/data_crawl/raw_data/gazette/'

    councillist = [
        'London Borough of Redbridge'

    ]
    for i in range(0, len(councillist)):
        infoPageName = datapath + councillist[i] + '.xlsx'
        task = Task5(infoPageName)
        task.run(councillist[i])
        # try: 
        #     infoPageName = datapath + councillist[i] + '.xlsx'
        #     url = urllist[i]
        #     task = Task5(infoPageName, chrome_path)
        #     task.run(url,councillist[i])
        # except Exception as exc:
        #     print(exc)
