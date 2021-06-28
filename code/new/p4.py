# coding: utf-8
import time
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.support import expected_conditions as Expect
from selenium.webdriver.support.ui import Select
from selenium.webdriver.support.wait import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support.ui import WebDriverWait as Wait
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.by import By
from selenium import webdriver
import requests
from bs4 import BeautifulSoup
import re
import tqdm
import random
import sys
import os
import xlwt
import io
import sys
import urllib.request
import openpyxl
from concurrent.futures import ThreadPoolExecutor
import concurrent.futures
import logging
import urllib3
import pandas as pd
import datetime
urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)
logging.basicConfig(level=logging.CRITICAL)

url = 'https://publicaccess.barnet.gov.uk/online-applications/applicationDetails.do?activeTab=contacts&keyVal=ZZZY6GJIXE811'



class Task5:
    def __init__(self, resultPath, chrome_path, linkPath):
        self.finalResult = []
        self.linkpath = linkPath
        self.resultPath = resultPath
        chrome_options = Options()
        self.line = 0
        self.xlsFile = openpyxl.Workbook()
        self.sheet1 = self.xlsFile.create_sheet(index=0)
        chrome_options.add_argument('--headless')
        self.max_workers = 32
        self.executor = ThreadPoolExecutor(max_workers=self.max_workers)

    def read_all_links(self, linkpath):
        links =pd.read_csv(linkpath, index_col=None, header=None, engine='python' ,sep = '\t', error_bad_lines=False).values
        return links


    def run(self, url, council):
        header = ['Council', 'Ref. No', 'Name', 'Address', 'Received',
                  'Validated', 'Status', 'MetaInfo', 'Link',
                  'Reference', 'Alternative Reference', 'Application Received',
                  'Application Validated', 'Address', 'Proposal', 'Status',
                  'Decision', 'Decision Issued Date', 'Appeal Status', 'Appeal Decision',
                  'Application Type', 'Decision', 'Actual Decision Level',
                  'Expected Decision Level', 'Case Officer', 'Parish', 'Ward',
                  'District Reference', 'Applicant Name', 'Applicant Address',
                  'Environmental Assessment Requested', 'Agent Name',
                  'Agent Company Name', 'Agent Address', 'Agent Phone Number',
                  'Agent Contacts', 'Councillors',
                  'Application Received Date', 'Application Validated Date',
                  'Expiry Date', 'Actual Committee Date', 'Latest Neighbour Consultation Date',
                  'Neighbour Consultation Expiry Date', 'Standard Consultation Date',
                  'Standard Consultation Expiry Date', 'Last Advertised In Press Date',
                  'Latest Advertisement Expiry Date', 'Last Site Notice Posted Date',
                  'Latest Site Notice Expiry Date', 'Statutory Expiry Date',
                  'Agreed Expiry Date', 'Decision Made Date', 'Decision Issued Date',
                  'Permission Expiry Date', 'Decision Printed Date',
                  'Environmental Impact Assessment Received', 'Temporary Permission Expiry Date', 'Internal Target Date', 'Determination Deadline', 'Target Date'
                  ]
        
        for col in range(1, len(header)+1):
            self.sheet1.cell(1, col).value = header[col-1]
        
        
        urls = self.read_all_links(self.linkpath)
        start = []
        end = []
        par = tqdm.tqdm(total = len(urls), ncols=100)
        for i in range(0, len(urls), self.max_workers):
            start.append(i)
            if i + self.max_workers >= len(urls):
                end.append(len(urls))
            else:
                end.append(i+self.max_workers)
        for i, s in enumerate(start):
            try: 
                par.update(self.max_workers)
                self.asyn_page(
                    url_list=urls[start[i]: end[i]], council=council)
            except Exception as exc:
                print(exc)
                continue   
        par.close()
        par = tqdm.tqdm(total = len(self.finalResult), ncols=100)
        for data in self.finalResult:
            par.update(1)
            for col in range(1, len(data)+1):
                self.sheet1.cell(self.line+2, col).value = data[col-1]            
            self.line += 1
        self.xlsFile.save(self.resultPath)
        par.close()
                
        self.xlsFile.close()
        print('total page is {}'.format(len(self.finalResult)))
        print('get webpage finish!')

    def __del__(self):       
        print('>>>>[Well Done]')

    def get_information(self, result, council):
        refNo = result[0]
        name = result[1]
        address = result[2]
        Received    = result[3]         
        Validated = result[4]
        Status = result[5]
        metaInfo = result[6]
        link = result[7]
        summary = []
        details = []
        contacts = []
        dates = []
        try:
            summarylink = link
            detailslink = link.replace('activeTab=summary', 'activeTab=details')
            contactslink = link.replace('activeTab=summary', 'activeTab=contacts')
            dateslink = link.replace('activeTab=summary', 'activeTab=dates')
            summary = self.get_summary(summarylink)
            details = self.get_details(detailslink)
            contacts = self.get_contacts( contactslink)
            dates = self.get_dates(dateslink)
        except Exception:
            Exception
        res = [council,refNo, name, address, Received,
                Validated, Status, metaInfo, link] + summary + details + contacts + dates
        return res
    

    def get_summary(self, url):
        summary = []
        page = requests.get(url,verify=False).text
        soup = BeautifulSoup(page, 'html5lib')
        keys = soup.select('#simpleDetailsTable > tbody > tr > th')
        values = soup.select('#simpleDetailsTable > tbody > tr > td')
        Reference = ''
        Alternative_Reference = ''
        Application_Received = ''
        Application_Validated = ''
        Address = ''
        Proposal = ''
        Status = ''
        Decision = ''
        Decision_Issued_Date = ''
        Appeal_Status = ''
        Appeal_Decision = ''

        
        for i in range(0,len(keys)):
            key = keys[i].get_text().strip()
            content = values[i].get_text().strip()
            if key == 'Reference':
                Reference = content
            elif key == 'Alternative Reference':
                Alternative_Reference = content
            elif key == 'Application Received':
                Application_Received = content
            elif key == 'Application Validated':
                Application_Validated = content
            elif key == 'Address':
                Address = content
            elif key == 'Proposal':
                Proposal = content
            elif key == 'Status':
                Status = content
            elif key == 'Decision':
                Decision = content
            elif key == 'Decision Issued Date':
                Decision_Issued_Date = content
            elif key == 'Appeal Status':
                Appeal_Status = content
            elif key == 'Appeal Decision':
                Appeal_Decision = content
        summary = [Reference, Alternative_Reference, Application_Received, Application_Validated,
                    Address, Proposal, Status, Decision, Decision_Issued_Date, Appeal_Status, Appeal_Decision]
        return summary

    def get_details(self, url):
        details = []
        page = requests.get(url,verify=False).text
        soup = BeautifulSoup(page, 'html5lib')
        keys = soup.select(' tbody > tr > th')
        values = soup.select('tbody > tr > td')
        Application_Type = ''
        Decision = ''
        Actual_Decision_Level = ''
        Expected_Decision_Level = ''
        Case_Officer = ''
        Parish = ''
        Ward = ''
        District_Reference = ''
        Applicant_Name = ''
        Applicant_Address = ''
        Environmental_Assessment_Requested = ''
        Agent_Name = ''
        Agent_Company_Name = ''
        Agent_Address = ''
        Agent_Phone_Number = ''
        for i in range(0,len(keys)):
            key = keys[i].get_text().strip()        
            content = values[i].get_text().strip()
            if key == 'Application Type':
                Application_Type = content
            elif key == 'Decision':
                Decision = content
            elif key == 'Actual Decision Level':
                Actual_Decision_Level = content
            elif key == 'Expected Decision Level':
                Expected_Decision_Level = content
            elif key == 'Case Officer':
                Case_Officer = content
            elif key == 'Parish':
                Parish = content
            elif key == 'Ward':
                Ward = content
            elif key == 'District Reference':
                District_Reference = content
            elif key == 'Applicant Name':
                Applicant_Name = content
            elif key == 'Applicant Address':
                Applicant_Address = content
            elif key == 'Environmental Assessment Requested':
                Environmental_Assessment_Requested = content
            elif key == 'Agent Name':
                Agent_Name = content
            elif key == 'Agent Company Name':
                Agent_Company_Name = content
            elif key == 'Agent Address':
                Agent_Address = content
            elif key == 'Agent Phone Number':
                Agent_Phone_Number = content
        details = [Application_Type, Decision, Actual_Decision_Level, Expected_Decision_Level, Case_Officer, Parish, Ward, District_Reference,
                        Applicant_Name, Applicant_Address, Environmental_Assessment_Requested, Agent_Name, Agent_Company_Name, Agent_Address, Agent_Phone_Number]
        return details

    def get_contacts(self, url):
        contacts = []
        page = requests.get(url,verify=False).text
        soup = BeautifulSoup(page, 'html5lib')
        agentsTable = soup.select(
                '#pa > div:nth-child(3) > div.content > div.tabcontainer > div.agents')
        councillorsTable = soup.select(
                '#pa > div:nth-child(3) > div.content > div.tabcontainer > div.councillors')
        agents = ''
        councillors = ''
        for agt in agentsTable:
            agents += agt.get_text() + '\n'

        for cot in councillorsTable:
            councillors += cot.get_text() + '\n'
        contacts = [agents, councillors]
        return contacts

    def get_dates(self, url):
        dates = []
        page = requests.get(url,verify=False).text
        soup = BeautifulSoup(page, 'html5lib')
        keys = soup.select('#simpleDetailsTable > tbody > tr > th')
        values = soup.select('#simpleDetailsTable > tbody > tr > td')
        Application_Received_Date = ''
        Application_Validated_Date = ''
        Expiry_Date = ''
        Actual_Committee_Date = ''
        Latest_Neighbour_Consultation_Date = ''
        Neighbour_Consultation_Expiry_Date = ''
        Standard_Consultation_Date = ''
        Standard_Consultation_Expiry_Date = ''
        Last_Advertised_In_Press_Date = ''
        Latest_Advertisement_Expiry_Date = ''
        Last_Site_Notice_Posted_Date = ''
        Latest_Site_Notice_Expiry_Date = ''
        Statutory_Expiry_Date = ''
        Agreed_Expiry_Date = ''
        Decision_Made_Date = ''
        Decision_Issued_Date = ''
        Permission_Expiry_Date = ''
        Decision_Printed_Date = ''
        Environmental_Impact_Assessment_Received = ''
        Temporary_Permission_Expiry_Date = ''
        Internal_Target_Date = ''
        Determination_Deadline = ''
        Target_Date = ''

        for i in range(0,len(keys)):
            key = keys[i].get_text().strip()        
            content = values[i].get_text().strip()
            if key == 'Application Received Date':
                Application_Received_Date = content
            elif key == 'Application Validated Date':
                Application_Validated_Date = content
            elif key == 'Expiry Date':
                Expiry_Date = content
            elif key == 'Actual Committee Date':
                Actual_Committee_Date = content
            elif key == 'Latest Neighbour Consultation Date':
                Latest_Neighbour_Consultation_Date = content
            elif key == 'Neighbour Consultation Expiry Date':
                Neighbour_Consultation_Expiry_Date = content
            elif key == 'Standard Consultation Date':
                Standard_Consultation_Date = content
            elif key == 'Standard Consultation Expiry Date':
                Standard_Consultation_Expiry_Date = content
            elif key == 'Last Advertised In Press Date':
                Last_Advertised_In_Press_Date = content
            elif key == 'Latest Advertisement Expiry Date':
                Latest_Advertisement_Expiry_Date = content
            elif key == 'Last Site Notice Posted Date':
                Last_Site_Notice_Posted_Date = content
            elif key == 'Latest Site Notice Expiry Date':
                Latest_Site_Notice_Expiry_Date = content
            elif key == 'Statutory Expiry Date':
                Statutory_Expiry_Date = content
            elif key == 'Agreed Expiry Date':
                Agreed_Expiry_Date = content
            elif key == 'Decision Made Date':
                Decision_Made_Date = content
            elif key == 'Decision Issued Date':
                Decision_Issued_Date = content
            elif key == 'Permission Expiry Date':
                Permission_Expiry_Date = content
            elif key == 'Decision Printed Date':
                Decision_Printed_Date = content
            elif key == 'Environmental Impact Assessment Received':
                Environmental_Impact_Assessment_Received = content
            elif key == 'Temporary Permission Expiry Date':
                Temporary_Permission_Expiry_Date = content
            elif key == 'Internal Target Date':
                Internal_Target_Date = content
            elif key == 'Determination Deadline':
                Determination_Deadline = content
            elif key == 'Target Date':
                Target_Date = content
        dates = [Application_Received_Date, Application_Validated_Date, Expiry_Date, Actual_Committee_Date,
                        Latest_Neighbour_Consultation_Date, Neighbour_Consultation_Expiry_Date, Standard_Consultation_Date,
                        Standard_Consultation_Expiry_Date, Last_Advertised_In_Press_Date, Latest_Advertisement_Expiry_Date,
                        Last_Site_Notice_Posted_Date, Latest_Site_Notice_Expiry_Date, Statutory_Expiry_Date,
                        Agreed_Expiry_Date, Decision_Made_Date, Decision_Issued_Date, Permission_Expiry_Date,
                        Decision_Printed_Date, Environmental_Impact_Assessment_Received, Temporary_Permission_Expiry_Date, 
                        Internal_Target_Date, Determination_Deadline, Target_Date]
        return dates

    def asyn_page(self, url_list, council):
        future_to_url  = dict()
        for i, url in enumerate(url_list):
            try: 
                t = self.executor.submit(self.get_information,
                                    result=url_list[i], council = council)
                future_to_url[t] = url    
            except Exception as exc:
                try: 
                    t = self.executor.submit(self.get_information,
                                        result=url_list[i], council = council)
                    future_to_url[t] = url
                except Exception as exc:
                    continue    
        for future in concurrent.futures.as_completed(future_to_url):
            url = future_to_url[future]
            try:
                data = future.result()
                if data != None:
                    self.finalResult.append(data)
            except Exception as exc:
                print('%r generated an exception: %s' % (url, exc))

if __name__ == '__main__':
    # chage the data path here
    datapath = 'D:/git/data_crawl/raw_data/gazette/'
    chrome_path = 'C:/Program Files/Google/Chrome/Application/chromedriver.exe'
    urllist = [
        'https://publicaccess.barnet.gov.uk/online-applications/search.do?action=advanced',
       
    ]
    councillist = [
        
        'London_Borough_of_Barnet', # can not connected
       
    ]
    for i in range(0, len(urllist)):        
        try: 
            infoPageName = datapath + councillist[i] + '2.xlsx'
            linkpath = datapath + councillist[i] +'link.tsv'
            url = urllist[i]
            task = Task5(infoPageName, chrome_path, linkpath)
            task.run(url,councillist[i])
        except Exception as exc:
            print(exc)
            continue
