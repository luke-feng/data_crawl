# coding: utf-8

import time
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.support import expected_conditions as Expect
from selenium.webdriver.support.ui import Select
from selenium.webdriver.support.wait import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support.ui import WebDriverWait as Wait
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.by import By
from selenium import webdriver
import requests
from bs4 import BeautifulSoup
import re
import tqdm
import random
import sys
import os
import xlwt
import io
import sys
import urllib.request
import openpyxl
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf8')

summarykey = []
detailskey = []
dateskey = []

def get_all_item(resultPath):
    chrome_options = Options()
    chrome_options.add_argument('--headless')
    browser = webdriver.Chrome(
        executable_path='C:/Program Files/Google/Chrome/Application/chromedriver.exe', options=chrome_options)
    browser1 = webdriver.Chrome(
        executable_path='C:/Program Files/Google/Chrome/Application/chromedriver.exe', options=chrome_options)
    url = 'https://www1.somersetwestandtaunton.gov.uk/online-applications/search.do?action=advanced'
    startday = ['01/01', '01/02', '01/03', '01/04', '01/05', '01/06',
                '01/07', '01/08', '01/09', '01/10', '01/11', '01/12']
    endday = ['01/02', '01/03', '01/04', '01/05', '01/06', '01/07',
              '01/08', '01/09', '01/10', '01/11', '01/12', '31/12']

    header = ['Ref. No', 'Name', 'Address', 'Received',
              'Validated', 'Status', 'MetaInfo', 'Link', 'Reference',
              'Alternative Reference', 'Application Received',
              'Application Validated', 'Address', 'Proposal',
              'Status', 'Appeal Status', 'Appeal Decision', 'Application Type'
              'Expected Decision Level',
              'Case Officer',
              'Parish',
              'Ward',
              'District Reference',
              'Applicant Name',
              'Agent Name',
              'Agent Company Name',
              'Agent Address',
              'Agent Phone Number',
              'Environmental Assessment Requested',
              'Contacts',
              'Application Received Date',
              'Application Validated Date',
              'Expiry Date',
              'Actual Committee Date',
              'Latest Neighbour Consultation Date',
              'Neighbour Consultation Expiry Date',
              'Standard Consultation Date',
              'Standard Consultation Expiry Date'	,
              'Last Advertised In Press Date',
              'Latest Advertisement Expiry Date',
              'Last Site Notice Posted Date',
              'Latest Site Notice Expiry Date',
              'Statutory Expiry Date',
              'Agreed Expiry Date',
              'Decision Made Date',
              'Permission Expiry Date',
              'Decision Printed Date',
              'Environmental Impact Assessment Received',
              'Temporary Permission Expiry Date']

    outputFile = resultPath
    xlsFile = openpyxl.Workbook()
    sheet1 = xlsFile.create_sheet(index=0)
    for col in range(1, len(header)+1):
        sheet1.cell(1, col).value = header[col-1]
    line = 2
    
    for year in range(2000, 2001):
        for mon in range(0, 12):
            qstart = startday[mon]+'/'+str(year)
            qend = endday[mon]+'/'+str(year)
            print(qstart, qend)
            browser.get(url)
            Wait(browser, 600).until(
                Expect.presence_of_element_located(
                    (By.CSS_SELECTOR, 'body>#idox>#pa>.container>.content>.tabcontainer>#advancedSearchForm>#dates>fieldset>.row>.col-dateFrom'))
            )
            fromdate = browser.find_element_by_id('applicationReceivedStart')
            fromdate.send_keys(qstart)
            enddate = browser.find_element_by_id('applicationReceivedEnd')
            enddate.send_keys(qend)
            browser.find_element_by_css_selector(
                'body>#idox>#pa>.container>.content>.tabcontainer>#advancedSearchForm>.buttons>.button.primary').click()
            browser.get(browser.current_url)
            Wait(browser, 600).until(
                Expect.presence_of_element_located(
                    (By.ID, 'resultsPerPage'))
            )
            resultsPerPage = browser.find_element_by_id('resultsPerPage')
            s1 = Select(resultsPerPage)
            s1.select_by_value('100')
            browser.find_element_by_css_selector(
                'body>#idox>#pa>.container>.content>#searchfilters>#searchResults>.button.primary').click()
            # browser.get(browser.current_url)
            searchResults = browser.find_elements_by_class_name('searchresult')

            while True:
                next = browser.find_elements_by_class_name('next')
                if len(next) > 0:
                    par = tqdm.tqdm( total=len(searchResults), ncols=100)
                    for result in searchResults:
                        par.update(1)
                        link = result.find_element_by_css_selector(
                            'a').get_attribute('href')
                        name = result.find_element_by_css_selector('a').text
                        address = result.find_element_by_css_selector(
                            '.address').text
                        metaInfo = result.find_element_by_css_selector(
                            '.metaInfo').text
                        ml = metaInfo.split('|')
                        refNo = ml[0].split(':')[1]
                        Received = ml[1].split(':')[1]
                        Validated = ml[2].split(':')[1]
                        Status = ml[3].split(':')[1]
                        information = get_information(browser1, link)
                        result = [refNo, name, address, Received,
                                  Validated, Status, metaInfo, link] + information
                        for col in range(1, len(result)+1):
                            sheet1.cell(line, col).value = result[col-1]
                        line += 1
                    par.close()
                    next_url = next[0].get_attribute('href')
                    browser.get(next_url)
                    searchResults = browser.find_elements_by_class_name(
                        'searchresult')
                else:
                    par = tqdm.tqdm( total=len(searchResults), ncols=100)
                    for result in searchResults:
                        par.update(1)
                        link = result.find_element_by_css_selector(
                            'a').get_attribute('href')
                        name = result.find_element_by_css_selector('a').text
                        address = result.find_element_by_css_selector(
                            '.address').text
                        metaInfo = result.find_element_by_css_selector(
                            '.metaInfo').text
                        ml = metaInfo.split('|')
                        refNo = ml[0].split(':')[1]
                        Received = ml[1].split(':')[1]
                        Validated = ml[2].split(':')[1]
                        Status = ml[3].split(':')[1]
                        information = get_information(browser1, link)
                        result = [refNo, name, address, Received,
                                  Validated, Status, metaInfo, link] + information
                        for col in range(1, len(result)+1):
                            sheet1.cell(line, col).value = result[col-1]
                        line += 1
                    par.close()
                    break
    xlsFile.save(outputFile)
    print('total page is {}'.format(str(line-1)))
    browser.close()
    browser1.close()
    print('get webpage finish!')

def get_information(browser, url):
    print(url)
    summary = []
    details = []
    contacts = []
    dates = []
    results = []
    summarylink = url
    detailslink = url.replace('activeTab=summary', 'activeTab=details')
    contactslink = url.replace('activeTab=summary', 'activeTab=contacts')
    dateslink = url.replace('activeTab=summary', 'activeTab=dates')
    summary = get_summary(browser,summarylink)
    details = get_details(browser,detailslink)
    contacts = get_contacts(browser,contactslink)
    dates = get_dates(browser,dateslink)
    results = summary + details + contacts + dates
    return results


def get_summary(browser, url):
    summary = []
    browser.get(url)
    simpleDetailsTable = browser.find_elements_by_id('simpleDetailsTable')
    if len(simpleDetailsTable)>0:
        trs = simpleDetailsTable[0].find_elements_by_css_selector('tbody>tr')
        for tr in trs:
            content = tr.find_element_by_css_selector('td').text
            key = tr.find_element_by_css_selector('th').text
            if key not in summarykey:
                summarykey.append(key)
            summary.append(content)
    return summary


def get_details(browser,url):
    details = []
    browser.get(url)
    simpleDetailsTable = browser.find_elements_by_id('applicationDetails')
    if len(simpleDetailsTable)>0:
        trs = simpleDetailsTable[0].find_elements_by_css_selector('tbody>tr')
        for tr in trs:
            content = tr.find_element_by_css_selector('td').text
            key = tr.find_element_by_css_selector('th').text
            if key not in detailskey:
                detailskey.append(key)
            details.append(content)
    return details


def get_contacts(browser,url):
    contacts = []
    browser.get(url)
    simpleDetailsTable = browser.find_element_by_class_name('tabcontainer')
    lines = simpleDetailsTable.text
    contacts = [lines]
    return contacts


def get_dates(browser,url):
    dates = []
    browser.get(url)
    simpleDetailsTable = browser.find_elements_by_id('simpleDetailsTable')
    if len(simpleDetailsTable)>0:
        trs = simpleDetailsTable[0].find_elements_by_css_selector('tbody>tr')
        for tr in trs:
            content = tr.find_element_by_css_selector('td').text
            key = tr.find_element_by_css_selector('th').text
            if key not in dateskey:
                dateskey.append(key)
            dates.append(content)
    return dates


datapath = 'D:/git/data_crawl/raw_data/gazette/'
infoPageName = datapath + 'result1.xlsx'
get_all_item(infoPageName)
print(summarykey)
print(detailskey)
print(dateskey)