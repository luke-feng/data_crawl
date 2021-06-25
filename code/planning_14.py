# coding: utf-8

import time
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.support import expected_conditions as Expect
from selenium.webdriver.support import wait
from selenium.webdriver.support.ui import Select
from selenium.webdriver.support.wait import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support.ui import WebDriverWait as Wait
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.by import By
from selenium import webdriver
import requests
from bs4 import BeautifulSoup
import re
import tqdm
import random
import sys
import os
import xlwt
import io
import sys
import urllib.request
import openpyxl
from concurrent.futures import ThreadPoolExecutor
import concurrent.futures
import datetime
from selenium.webdriver.common.action_chains import ActionChains
import urllib3
urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)
import unicodedata

# chrome_options = Options()
# chrome_options.add_argument('--headless')
# chrome_options.add_argument("--log-level=OFF")
# chrome_options.add_experimental_option('excludeSwitches', ['enable-logging'])
# # Change your chrome path here
# chrome_path = 'C:/Program Files/Google/Chrome/Application/chromedriver.exe'
# driver = webdriver.Chrome(
#     executable_path=chrome_path, options=chrome_options, service_log_path='NUL')

# url = 'https://planning.islington.gov.uk/northgate/planningexplorer/generalsearch.aspx'
# driver.get(url)

# fromdate = driver.find_element_by_xpath('/html/body/form/div/div/div[4]/ul/li[7]/div/input')
# qstart = '24-12-2020'
# qend = '31-12-2020'
# fromdate.send_keys(qstart)
# enddate = driver.find_element_by_xpath('/html/body/form/div/div/div[4]/ul/li[8]/div/input')
# enddate.send_keys(qend)
# sbd = driver.find_element_by_xpath('/html/body/form/div/div/div[4]/ul/li[6]/div/input')
# sbd.click()
# search = driver.find_element_by_xpath('/html/body/form/div/div/div[4]/center/input')
# print(search.text)
# search.click()
# lens = driver.find_element_by_xpath('/html/body/form/div/div[1]/span[1]').text
# print(lens)

# trs = driver.find_elements_by_xpath('/html/body/form/div/table/tbody/tr')[1:]
# print(len(trs))
# driver.find_element_by_css_selector
# url = 'https://planning.islington.gov.uk/Northgate/PlanningExplorer/Generic/StdDetails.aspx?PT=Planning%20Applications%20On-Line&TYPE=PL/PlanningPK.xml&PARAM0=496297&XSLT=/Northgate/PlanningExplorer/SiteFiles/Skins/Islington/xslt/PL/PLDetails.xslt&FT=Planning%20Application%20Details&PUBLIC=Y&XMLSIDE=&DAURI=PLANNING'
# page = requests.get(url,verify=False).text

# soup = BeautifulSoup(page, 'lxml')
# kvs1 = soup.select('#Template > div:nth-child(7) > ul > li > div')
# kvs2 = soup.select('#Template > div:nth-child(9) > ul > li > div')
# kvs = kvs1 + kvs2
# for i in range(0,len(kvs)):
#     key = kvs[i].contents[1].get_text()
#     value = kvs[i].contents[2]


class Task5:
    def __init__(self, resultPath, chrome_path):
        self.finalResult = []
        self.resultPath = resultPath
        chrome_options = Options()
        self.line = 0
        self.v = []
        self.xlsFile = openpyxl.Workbook()
        self.sheet1 = self.xlsFile.create_sheet(index=0)
        chrome_options.add_argument('--headless')
        chrome_options.add_experimental_option ('excludeSwitches', ['enable-logging'])
        self.max_workers=10
        self.executor = ThreadPoolExecutor(max_workers=self.max_workers)
        # for main thread
        self.driver = webdriver.Chrome(
            executable_path=chrome_path, options=chrome_options)

    def get_search_results_page(self, url, qstart, qend, council):
        self.driver.get(url)        
        fromdate = self.driver.find_element_by_xpath('/html/body/div/div/div[2]/form/div/div/div[5]/ul/li[7]/div/input')
        fromdate.send_keys(qstart)
        enddate = self.driver.find_element_by_xpath('/html/body/div/div/div[2]/form/div/div/div[5]/ul/li[8]/div/input')
        enddate.send_keys(qend)
        sbd = self.driver.find_element_by_xpath('/html/body/div/div/div[2]/form/div/div/div[5]/ul/li[6]/div/input')
        sbd.click()
        search = self.driver.find_element_by_xpath('/html/body/div/div/div[2]/form/div/div/input')
        search.click()
        # searchResults = searchResults[0].find_elements_by_css_selector ('table > tbody > tr > td ')
        print(council, qstart, qend)



    def run(self, url, council):
        header = ['Council', 'Ref. No', 'Name', 'Address', 'Received',
                'Validated', 'Status', 'Decision', 'Development Description', 'Link',
                'Application_Registered', 'Comments_Until', 'Date_of_Committee', 'Decision',
                'Appeal_Lodged', 'Appeal_Decision', 'Application_Number', 'Site_Address',
                'Application_Type', 'Development_Type', 'Proposal', 'Current_Status', 
                'Applicant', 'Agent', 'Wards', 'Advertised',
                'Constituency', 'Location_Co_ordinates', 'Parishes', 'OS_Mapsheet',
                 'Appeal_Submitted', 'Appeal_Decision', 'Case_Officer', 'Division',
                'Planning_Officer','Recommendation', 'Determination_Level',
                'Received', 'First_Advertised','Registered',  'First_Site_Notice',
                'Valid', 'Consultation_Expiry', 'Validated', 'Stat_Cons_Expiry_Date',
                'Decision_Expiry', 'Date_of_First_Consultation', 'Extended_Expiry'
                ]

        for col in range(1, len(header)+1):
            self.sheet1.cell(1, col).value = header[col-1]

        qstart = datetime.date(2000,1,1)
        last = datetime.date(2020,12,31)
        qend = qstart
        while qend < last:
            qend = qstart + datetime.timedelta(days=20)
            if qend >= last:
                qend = last
            # par = tqdm.tqdm(ncols=100)
            sstart = qstart.strftime("%d-%m-%Y")
            send = qend.strftime("%d-%m-%Y")
            qstart = qstart + datetime.timedelta(days=21)       
            try: 
                self.get_search_results_page(url, sstart, send, council)
                
                totn = self.driver.find_element_by_xpath('/html/body/div/div/div[2]/form/div[3]/div[1]/span').text.split('of')[1]
                par = tqdm.tqdm(total=int(totn), ncols=100)
                while True:
                    next = self.driver.find_elements_by_xpath('/html/body/div/div/div[2]/form/div[3]/div[3]/a')
                    ntitle = 'Go to next page'
                    bnext = 0
                    searchResults = self.driver.find_elements_by_xpath('/html/body/div/div/div[2]/form/div[3]/table/tbody/tr')[1:]
                    start = []
                    end = []  
                    
                    for i in range(0, len(searchResults), self.max_workers):
                        start.append(i)
                        if i + self.max_workers >= len(searchResults):
                            end.append(len(searchResults))
                        else:
                            end.append(i+self.max_workers)
                    for i, s in enumerate(start):
                        self.asyn_page(
                            url_list=searchResults[start[i]:end[i]], council=council)
                        par.update(self.max_workers)
                    if len(next) > 0:
                        for n in next:
                            na = n.find_elements_by_css_selector('img')
                            if len(na) > 0:
                                title = na[0].get_attribute('title')
                                                    
                                if ntitle in title:
                                    new_link = n.get_attribute('href')
                                
                                    self.driver.get(new_link)
                                    sss = self.driver.find_element_by_xpath('/html/body/div/div/div[2]/form/div[3]/div[1]/span').text
                
                                    bnext += 1
                                    break
                        if bnext == 0:
                            break                            
                    else:
                        break
            except Exception as exc:
                print(sstart, send, exc)
                continue

        par = tqdm.tqdm(total = len(self.finalResult), ncols=100)
        for data in self.finalResult:
            par.update(1)
            for col in range(1, len(data)+1):
                self.sheet1.cell(self.line+2, col).value = data[col-1]            
            self.line += 1
        self.xlsFile.save(self.resultPath)
        par.close()
        self.xlsFile.close()
        print('total page is {}'.format(len(self.finalResult)))
        print('get webpage finish!')
    
    def __del__(self):
        self.driver.close()
        print('>>>>[Well Done]')

    def get_information(self, result, council):
        refNo = ''
        Received= ''
        Registered= ''
        Status= ''
        Decision = ''
        name = ''
        address = ''
        metaInfo = ''
        details = []
        tds = result.find_elements_by_css_selector('td')

        if len(tds)>0:
            refNo = tds[0].find_element_by_css_selector('a').text
            link = tds[0].find_element_by_css_selector('a').get_attribute('href')
            address = tds[1].text
            metaInfo = tds[2].text
            Registered = tds[4].text
            Status = tds[3].text   
            Decision = tds[5].text

            details = self.get_details(link)

            res = [council,refNo, name, address, Received,
                Registered, Status, Decision,  metaInfo, link] + details
            return res
        else:
            return None
        
    def get_details(self, url):
        details = []
        dateslink = ''
        page = requests.get(url,verify=False).text
        soup = BeautifulSoup(page, 'lxml')
        soup.prettify(formatter=self.remove_chara)
        kvs1 = soup.select('#Template > div:nth-child(7) > ul > li > div')
        kvs2 = soup.select('#Template > div:nth-child(9) > ul > li > div')
        kvs = kvs1 + kvs2

        Application_Registered = ''
        Comments_Until = ''
        Date_of_Committee = ''
        Decision = ''
        Appeal_Lodged = ''
        Appeal_Decision = ''
        Application_Number = ''
        Site_Address = ''
        Application_Type = ''
        Development_Type = ''
        Proposal= ''
        Current_Status = ''
        Applicant = ''
        Agent = ''
        Wards = ''
        Advertised = ''
        Constituency = ''
        Location_Co_ordinates = ''
        Parishes = ''
        OS_Mapsheet = ''
        Appeal_Submitted = ''
        Appeal_Decision = ''
        Case_Officer = ''
        Division= ''
        Planning_Officer = ''
        Recommendation = ''
        Determination_Level = ''

        for i in range(0,len(kvs)):
            key = kvs[i].contents[1].get_text().strip()
            value = kvs[i].contents[2].strip()
            if key not in self.v:
                self.v.append(key)
            if key == 'Application Registered':
                Application_Registered = value
            elif key == 'Comments Until':
                Comments_Until = value
            elif key == 'Date of Committee':
                Date_of_Committee = value
            elif key == 'Decision':
                Decision = value
            elif key == 'Appeal Lodged':
                Appeal_Lodged = value
            elif key == 'Appeal Decision':
                Appeal_Decision = value
            elif key == 'Application Number':
                Application_Number = value
            elif key == 'Site Address':
                Site_Address = value
            elif key == 'Application Type':
                Application_Type = value
            elif key == 'Development Type':
                Development_Type = value
            elif key == 'Proposal':
                Proposal = value
            elif key == 'Current Status':
                Current_Status = value
            elif key == 'Applicant':
                Applicant = value
            elif key == 'Agent':
                Agent = value
            elif key == 'Wards':
                Wards = value
            elif key == 'Advertised':                
                Advertised = value
            elif key == 'Constituency':
                Constituency = value
            elif key == 'Location Co ordinates':
                Location_Co_ordinates = value
            elif key == 'Parishes':
                Parishes = value
            elif key == 'OS Mapsheet':
                OS_Mapsheet = value
            elif key == 'Appeal Submitted?':
                Appeal_Submitted = value
            elif key == 'Appeal Decision':
                Appeal_Decision = value
            elif key == 'Case Officer / Tel':
                Case_Officer = value
            elif key == 'Division':
                Division = value
            elif key == 'Planning Officer':
                Planning_Officer = value
            elif key == 'Recommendation':
                Recommendation = value
            elif key == 'Determination Level':
                Determination_Level = value

        dates = []
        kvs3 = soup.select('#Template > div:nth-child(11) > ul > li:nth-child(1) > div')
        link = kvs3[0].find('a')
        link = link['href']
        a = link.split()
        dateslink = 'https://planning.islington.gov.uk/Northgate/PlanningExplorer/Generic/'
        for ass in a:
            dateslink += ass
        dates = self.get_detes(dateslink)

        details = [Application_Registered, Comments_Until, Date_of_Committee, Decision,
        Appeal_Lodged, Appeal_Decision, Application_Number, Site_Address,
        Application_Type, Development_Type, Proposal, Current_Status, 
        Applicant, Agent, Wards, Advertised,
        Constituency, Location_Co_ordinates, Parishes, OS_Mapsheet,
        Appeal_Submitted, Appeal_Decision, Case_Officer, Division,
        Planning_Officer,Recommendation, Determination_Level
        ]
        details = details  + dates
        return details

    def remove_chara(self, markup):
        return markup.replace("&nbsp;","")

    def get_detes(self, url):
        dates = []
        page = requests.get(url,verify=False).text
        soup = BeautifulSoup(page, 'html5lib')
        soup.prettify(formatter=self.remove_chara)
        kvs1 = soup.select('#Template > div:nth-child(8) > ul > li > div')
        kvs = kvs1 
        Received = ''
        First_Advertised = ''
        Registered = ''
        First_Site_Notice = ''
        Valid = ''
        Consultation_Expiry = ''
        Validated = ''
        Stat_Cons_Expiry_Date = ''
        Decision_Expiry = ''
        Date_of_First_Consultation = ''
        Extended_Expiry = ''
        

        for i in range(0,len(kvs)):
            key = kvs[i].contents[1].get_text().strip()
            value = kvs[i].contents[2].strip()
            if key not in self.v:
                self.v.append(key)
            if key == 'Received':
                Received = value
            elif key == 'First Advertised':
                First_Advertised = value
            elif key == 'Registered':
                Registered = value
            elif key == 'First Site Notice':
                First_Site_Notice = value
            elif key == 'Valid':
                Valid = value
            elif key == 'Consultation Expiry':
                Consultation_Expiry = value
            elif key == 'Validated':
                Validated = value
            elif key == 'Stat Cons Expiry Date':
                Stat_Cons_Expiry_Date = value
            elif key == 'Decision Expiry':
                Decision_Expiry = value
            elif key == 'Date of First Consultation':
                Date_of_First_Consultation = value
            elif key == 'Extended Expiry':
                Extended_Expiry = value

        dates = [Received, First_Advertised,Registered,  First_Site_Notice,
            Valid, Consultation_Expiry, Validated, Stat_Cons_Expiry_Date,
            Decision_Expiry, Date_of_First_Consultation, Extended_Expiry
             ]


        return dates

    def asyn_page(self, url_list, council):
        future_to_url  = dict()
        for i, url in enumerate(url_list):
            t = self.executor.submit(self.get_information,
                                result=url_list[i], council = council)
            future_to_url[t] = url               
        for future in concurrent.futures.as_completed(future_to_url):
            url = future_to_url[future]
            try:
                data = future.result()
                if data != None:
                    self.finalResult.append(data)
            except Exception as exc:
                print('%r generated an exception: %s' % (url, exc))


if __name__ == '__main__':
    # chage the data path here
    datapath = 'D:/git/data_crawl/raw_data/gazette/'
    chrome_path = 'C:/Program Files/Google/Chrome/Application/chromedriver.exe'
    urllist = [
        'https://planning.islington.gov.uk/northgate/planningexplorer/generalsearch.aspx'
    ]
    councillist = [
        'London Borough of Islington'
        ]
    for i in range(0, len(urllist)):
        infoPageName = datapath + councillist[i] + '.xlsx'
        url = urllist[i]
        task = Task5(infoPageName, chrome_path)
        task.run(url,councillist[i])
        # print(task.v)